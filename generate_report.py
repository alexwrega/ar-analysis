#!/usr/bin/env python3
"""
Alpha Read Analysis Report Generator
Analyzes reading plans for grades 3-12 and generates a self-contained HTML report.
"""

import json
import os
import re
from collections import defaultdict, Counter
from html import escape

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ============================================================================
# CONFIGURATION
# ============================================================================

GRADE_DATA_PATHS = {
    3: "/Users/alexandra/Claude Code/Alpha Read Viewer/data/grade3.json",
    4: "/Users/alexandra/Claude Code/Alpha Read Viewer/data/grade4.json",
    5: "/Users/alexandra/Claude Code/Alpha Read Viewer/data/grade5.json",
    6: "/Users/alexandra/Claude Code/Alpha Read Viewer/data/grade6.json",
    7: "/Users/alexandra/Claude Code/Alpha Read Viewer/data/grade7.json",
    8: "/Users/alexandra/Claude Code/Alpha Read Viewer/data/grade8.json",
    9: "/Users/alexandra/Claude Code/Alpha Read Viewer 9-12/data/grade9.json",
    10: "/Users/alexandra/Claude Code/Alpha Read Viewer 9-12/data/grade10.json",
    11: "/Users/alexandra/Claude Code/Alpha Read Viewer 9-12/data/grade11.json",
    12: "/Users/alexandra/Claude Code/Alpha Read Viewer 9-12/data/grade12.json",
}

OUTPUT_PATH = "/Users/alexandra/Claude Code/Alpha Read Analysis/report.html"

# ============================================================================
# FLAGGED ITEMS FROM APPROPRIATENESS AUDIT
# (Source: https://ilmych.github.io/reading-appropriateness/)
# ============================================================================

FLAGGED_ITEMS = [
    # HIGH severity
    {"title": "Bully by Radiotopia/PRX", "status": "Used", "flags": "Profanity (moderate), Dark themes (severe), Violence (severe)", "severity": "HIGH"},
    {"title": "An Ordinary Man by Paul Rusesabagina (Part I)", "status": "Unused", "flags": "Violence (severe), Dark themes (severe)", "severity": "HIGH"},
    {"title": "The Lottery by Shirley Jackson (Part II)", "status": "Used", "flags": "Violence (severe), Dark themes (severe)", "severity": "HIGH"},
    {"title": "The American Embassy by Chimamanda Ngozi Adichie (Part I)", "status": "Used", "flags": "Dark themes (severe), Violence (severe), Sexual (mild)", "severity": "HIGH"},
    {"title": "Balboa by Sabina Murray", "status": "Unused", "flags": "Profanity (soft), Violence (severe), Dark themes (severe)", "severity": "HIGH"},
    {"title": "from Night by Elie Wiesel", "status": "Used", "flags": "Profanity (soft), Dark themes (severe), Violence (moderate)", "severity": "HIGH"},
    {"title": "Once Upon a Time by Nadine Gordimer", "status": "Used", "flags": "Dark themes (severe), Violence (moderate)", "severity": "HIGH"},
    {"title": "from Maus by Art Spiegelman", "status": "Unused", "flags": "Dark themes (severe), Violence (moderate)", "severity": "HIGH"},
    {"title": "The Fall of the House of Usher by Edgar Allan Poe (Part IV)", "status": "Unused", "flags": "Dark themes (severe), Violence (moderate)", "severity": "HIGH"},
    {"title": "from The Jungle by Upton Sinclair", "status": "Unused", "flags": "Drug indicators, Violence (moderate), Dark themes (severe)", "severity": "HIGH"},
    {"title": "The Crucible, Act I by Arthur Miller (Part IV)", "status": "Unused", "flags": "Profanity (soft), Dark themes (moderate), Sexual (severe)", "severity": "HIGH"},
    {"title": "from Farewell to Manzanar", "status": "Unused", "flags": "Drug indicators, Dark themes (moderate), Substance (severe)", "severity": "HIGH"},
    {"title": "Ambush by Tim O'Brien", "status": "Used", "flags": "Dark themes (moderate), Violence (severe)", "severity": "HIGH"},
    {"title": "A Rose for Emily by William Faulkner (Part II)", "status": "Unused", "flags": "Dark themes (severe), Violence (moderate), Social issues", "severity": "HIGH"},
    {"title": "Cask of Amontillado by Edgar Allan Poe", "status": "Used", "flags": "Dark themes (severe), Substance (moderate), Violence (moderate)", "severity": "HIGH"},
    {"title": "Dulce et Decorum Est by Wilfred Owen", "status": "Used", "flags": "Drug indicators, Violence (severe), Dark themes (moderate)", "severity": "HIGH"},
    {"title": "The Wife of Bath's Tale by Geoffrey Chaucer (Part I)", "status": "Unused", "flags": "Profanity (moderate), Sexual (severe), Dark themes (moderate)", "severity": "HIGH"},
    {"title": "The Tragedy of Macbeth, Act II by William Shakespeare", "status": "Used", "flags": "Profanity (strong/soft), Violence (moderate), Dark themes (severe)", "severity": "HIGH"},
    {"title": "The Tragedy of Macbeth, Act IV by William Shakespeare", "status": "Used", "flags": "Profanity (soft), Violence (moderate), Dark themes (severe)", "severity": "HIGH"},
    {"title": "To His Coy Mistress by Andrew Marvell", "status": "Used", "flags": "Sexual (severe)", "severity": "HIGH"},
    {"title": "Shooting an Elephant by George Orwell (Part I)", "status": "Used", "flags": "Sexual indicators, Violence (severe), Dark themes (moderate)", "severity": "HIGH"},
    {"title": "A Village After Dark by Kazuo Ishiguro (Part II)", "status": "Used", "flags": "Sexual indicators, Sexual (severe)", "severity": "HIGH"},
    {"title": "from The 57 Bus by Dashka Slater (Part I)", "status": "Used", "flags": "Self-harm indicators, Dark themes (moderate), Violence (moderate)", "severity": "HIGH"},
    {"title": "The Prisoner Who Wore Glasses by Bessie Head", "status": "Used", "flags": "Profanity (soft), Self-harm indicators, Violence (moderate)", "severity": "HIGH"},
    {"title": "from The Odyssey: A Dramatic Retelling", "status": "Used", "flags": "Profanity (moderate/soft), Self-harm indicators", "severity": "HIGH"},
    {"title": "Entwined by Brian Tobin (Part III)", "status": "Unused", "flags": "Violence indicators, Self-harm indicators, Dark themes (moderate)", "severity": "HIGH"},
    {"title": "The Night Face Up by Julio Cortazar (Part II)", "status": "Used", "flags": "Violence indicators, Self-harm indicators, Sexual indicators", "severity": "HIGH"},
    {"title": "Why Seeing (the Unexpected) Is Often Not Believing", "status": "Used", "flags": "Profanity (strong), Violence (moderate), Dark themes (moderate)", "severity": "HIGH"},
    {"title": "Elsewhere by Derek Walcott", "status": "Used", "flags": "Self-harm indicators, Violence (moderate), Dark themes (moderate)", "severity": "HIGH"},
    {"title": "The Most Dangerous Game by Richard Connel (Part IV)", "status": "Unused", "flags": "Profanity (moderate), Self-harm indicators, Violence (moderate)", "severity": "HIGH"},
    {"title": "The Lowest Animal by Mark Twain", "status": "Unused", "flags": "Violence indicators, Sexual indicators, Dark themes (moderate)", "severity": "HIGH"},
    {"title": "The Crucible, Act IV by Arthur Miller (Part II)", "status": "Unused", "flags": "Profanity (soft), Self-harm indicators, Violence (moderate)", "severity": "HIGH"},
    {"title": "An Occurrence at Owl Creek Bridge by Ambrose Bierce (Part I)", "status": "Unused", "flags": "Self-harm indicators, Violence (moderate), Dark themes (moderate)", "severity": "HIGH"},
    {"title": "from Beowulf by Burton Raffel (Part II)", "status": "Unused", "flags": "Profanity (soft), Violence indicators, Self-harm indicators", "severity": "HIGH"},
    {"title": "The Tragedy of Hamlet, Act IV Scenes 5-6", "status": "Unused", "flags": "Profanity (strong/soft), Dark themes (moderate)", "severity": "HIGH"},
    {"title": "from Writing as an Act of Hope by Isabel Allende", "status": "Used", "flags": "Violence indicators, Sexual indicators, Violence (moderate)", "severity": "HIGH"},
    {"title": "Life Is a Free Gift", "status": "Used", "flags": "Self-harm indicators, Dark themes (moderate)", "severity": "HIGH"},
    {"title": "Romeo Is a Dirtbag by Lois Leveen", "status": "Unused", "flags": "Self-harm indicators, Sexual (moderate)", "severity": "HIGH"},
    {"title": "How the Children of Birmingham Changed the Civil-Rights Move", "status": "Used", "flags": "Profanity (strong), Violence (moderate)", "severity": "HIGH"},
    {"title": "The Tempest, Act I, Scene II by William Shakespeare", "status": "Unused", "flags": "Profanity (strong/soft), Violence indicators", "severity": "HIGH"},
    {"title": "Thomas Jefferson: The Best of Enemies by Ron Chernow", "status": "Unused", "flags": "Violence indicators, Sexual (moderate)", "severity": "HIGH"},
    {"title": "The Crucible, Act II by Arthur Miller (Part III)", "status": "Unused", "flags": "Profanity (strong), Dark themes (moderate)", "severity": "HIGH"},
    {"title": "The Pedestrian by Ray Bradbury", "status": "Used", "flags": "Profanity (strong), Dark themes (moderate)", "severity": "HIGH"},
    {"title": "The Prologue from The Canterbury Tales", "status": "Unused", "flags": "Profanity (strong/soft), Substance (moderate)", "severity": "HIGH"},
    {"title": "The Tragedy of Hamlet, Act I Scene 2", "status": "Unused", "flags": "Profanity (strong/soft), Dark themes (moderate)", "severity": "HIGH"},
    {"title": "The Lagoon by Joseph Conrad (Part II)", "status": "Unused", "flags": "Profanity (strong), Dark themes (moderate)", "severity": "HIGH"},
    {"title": "The Chimney Sweeper by William Blake", "status": "Used", "flags": "Profanity (strong), Sexual indicators, Dark themes (moderate)", "severity": "HIGH"},
    {"title": "The Tempest, Act II, Scene I", "status": "Unused", "flags": "Profanity (strong)", "severity": "HIGH"},
    {"title": "In La Rinconada, Peru, Searching for Beauty", "status": "Used", "flags": "Profanity (soft), Violence indicators, Drug indicators", "severity": "HIGH"},
    {"title": "The World on the Turtle's Back: Myth", "status": "Used", "flags": "Self-harm indicators", "severity": "HIGH"},
    {"title": "Morte d'Arthur by Alfred, Lord Tennyson", "status": "Used", "flags": "Profanity (strong)", "severity": "HIGH"},
    {"title": "The Ugly Duckling by Hans Christian Andersen (Part I)", "status": "Used", "flags": "Profanity (strong)", "severity": "HIGH"},
    {"title": "from Hag-Seed by Margaret Atwood", "status": "Used", "flags": "Profanity (strong/moderate/soft)", "severity": "HIGH"},
    {"title": "A White Heron by Sarah Orne Jewett (Part I)", "status": "Unused", "flags": "Profanity (strong)", "severity": "HIGH"},
    {"title": "What You Don't Know Can Kill You by Jason Daley (Part I)", "status": "Used", "flags": "Profanity (soft), Violence indicators, Drug indicators", "severity": "HIGH"},
    {"title": "The Tragedy of Hamlet, Act I Scene 1", "status": "Unused", "flags": "Profanity (strong)", "severity": "HIGH"},
    {"title": "Pavavu", "status": "Used", "flags": "Self-harm indicators", "severity": "HIGH"},
    {"title": "A Monument to Revolutionary Trans Activists", "status": "Used", "flags": "Profanity (soft), Self-harm indicators", "severity": "HIGH"},
    {"title": "from Dust Tracks on a Road by Zora Neale Hurston (Part II)", "status": "Unused", "flags": "Profanity (strong)", "severity": "HIGH"},
    {"title": "The Men in the Storm by Stephen Crane", "status": "Used", "flags": "Profanity (soft), Self-harm indicators", "severity": "HIGH"},
    {"title": "More Strange Adventures", "status": "Used", "flags": "Self-harm indicators", "severity": "HIGH"},
    {"title": "The Adventure of the Lions", "status": "Used", "flags": "Self-harm indicators", "severity": "HIGH"},
    {"title": "Act 2, Scene 1", "status": "Used", "flags": "Violence indicators", "severity": "HIGH"},
    {"title": "Act 5 Retold", "status": "Used", "flags": "Self-harm indicators", "severity": "HIGH"},
    {"title": "The Trap of the White Sphinx", "status": "Used", "flags": "Self-harm indicators", "severity": "HIGH"},
    {"title": "The Chimney Sweeper Poems by William Blake", "status": "Used", "flags": "Profanity (strong), Sexual indicators", "severity": "HIGH"},
    {"title": "The Next Targets", "status": "Used", "flags": "Self-harm indicators", "severity": "HIGH"},
    {"title": "Okinawa", "status": "Used", "flags": "Violence indicators, Self-harm indicators", "severity": "HIGH"},
    {"title": "About Mary Shelley", "status": "Used", "flags": "Self-harm indicators", "severity": "HIGH"},
    {"title": "Chapter 4", "status": "Used", "flags": "Violence indicators", "severity": "HIGH"},
    {"title": "Emmett Till: Memories of a Murder in Mississippi", "status": "Used", "flags": "Violence indicators", "severity": "HIGH"},
    {"title": "Truth at All Costs by Marie Colvin", "status": "Used", "flags": "Violence indicators", "severity": "HIGH"},
    {"title": "Neither Justice nor Forgetting", "status": "Used", "flags": "Violence indicators", "severity": "HIGH"},
    {"title": "Let South Africa Show the World How to Forgive", "status": "Used", "flags": "Violence indicators", "severity": "HIGH"},
    {"title": "A Dish Best Served Cold by Aminatta Forna", "status": "Used", "flags": "Violence indicators", "severity": "HIGH"},
    {"title": "The Tragedy of Macbeth, Act III, Scenes 4-6", "status": "Used", "flags": "Violence indicators", "severity": "HIGH"},
    {"title": "The Voyage of the James Caird", "status": "Unused", "flags": "Self-harm indicators", "severity": "HIGH"},
    {"title": "from Nature, from Self-Reliance by Ralph Waldo Emerson", "status": "Used", "flags": "Self-harm indicators", "severity": "HIGH"},
    {"title": "A Literature of Place by Barry Lopez", "status": "Unused", "flags": "Profanity (strong/soft)", "severity": "HIGH"},
    {"title": "What You Don't Know Can Kill You (Part II)", "status": "Used", "flags": "Violence indicators", "severity": "HIGH"},
    {"title": "Focus Period: 1950-Present", "status": "Used", "flags": "Violence indicators", "severity": "HIGH"},
    {"title": "Chapter V: Children on the Road", "status": "Used", "flags": "Profanity (strong)", "severity": "HIGH"},
    {"title": "The Tragedy of Hamlet, Act II Scene 2 (Part III)", "status": "Unused", "flags": "Profanity (moderate/soft), Violence indicators", "severity": "HIGH"},
    {"title": "Hamlet's Dull Revenge by Rene Girard", "status": "Unused", "flags": "Violence indicators", "severity": "HIGH"},
    {"title": "The Tragedy of Macbeth, Act III Scenes 4-6", "status": "Unused", "flags": "Violence indicators", "severity": "HIGH"},
    {"title": "Elegy Written in a Country Churchyard", "status": "Used", "flags": "Profanity (strong), Drug indicators", "severity": "HIGH"},
    {"title": "Marriage Is a Private Affair by Chinua Achebe", "status": "Used", "flags": "Violence indicators", "severity": "HIGH"},
    # MIDDLE severity
    {"title": "The Tempest, Act III by William Shakespeare", "status": "Unused", "flags": "Sexual indicators, Violence (moderate), Dark themes (moderate)", "severity": "MIDDLE"},
    {"title": "Lamb to the Slaughter by Roald Dahl (Part I)", "status": "Used", "flags": "Dark themes (moderate), Violence (moderate), Substance (moderate)", "severity": "MIDDLE"},
    {"title": "Pyramus and Thisbe by Ovid", "status": "Used", "flags": "Violence (moderate), Dark themes (moderate), Sexual (moderate)", "severity": "MIDDLE"},
    {"title": "Oedipus the King, Part III", "status": "Used", "flags": "Profanity (soft), Dark themes (moderate), Sexual (moderate)", "severity": "MIDDLE"},
    {"title": "The Tragedy of Macbeth, Act I, Scenes 4-7", "status": "Used", "flags": "Profanity (moderate/soft), Sexual indicators, Violence (moderate)", "severity": "MIDDLE"},
    {"title": "To My Old Master by Jourdon Anderson", "status": "Unused", "flags": "Dark themes (moderate), Sexual (moderate), Violence (moderate)", "severity": "MIDDLE"},
    {"title": "The Crucible, Act I by Arthur Miller (Part III)", "status": "Unused", "flags": "Profanity (moderate), Sexual indicators, Dark themes (moderate)", "severity": "MIDDLE"},
    {"title": "The Leap by Louise Erdrich (Part I)", "status": "Used", "flags": "Violence (moderate), Dark themes (moderate)", "severity": "MIDDLE"},
    {"title": "The Tragedy of Hamlet, Act III Scene 4", "status": "Unused", "flags": "Profanity (soft), Drug indicators, Violence (moderate)", "severity": "MIDDLE"},
    {"title": "The Tragedy of Hamlet, Act IV Scene 7", "status": "Unused", "flags": "Sexual indicators, Dark themes (moderate), Violence (moderate)", "severity": "MIDDLE"},
    {"title": "The Tragedy of Hamlet, Act V Scene 2 (Part II)", "status": "Unused", "flags": "Dark themes (moderate), Substance (moderate), Violence (moderate)", "severity": "MIDDLE"},
    {"title": "Education Protects Women from Abuse", "status": "Unused", "flags": "Sexual indicators, Sexual (moderate), Violence (moderate)", "severity": "MIDDLE"},
    {"title": "from The 57 Bus by Dashka Slater (Part II)", "status": "Used", "flags": "Sexual indicators, Violence (moderate), Dark themes (moderate)", "severity": "MIDDLE"},
    {"title": "from The 57 Bus by Dashka Slater (Part III)", "status": "Used", "flags": "Dark themes (moderate), Violence (moderate)", "severity": "MIDDLE"},
    {"title": "Unsolved Vigilante Murder in the Heartland", "status": "Used", "flags": "Violence (moderate), Dark themes (moderate)", "severity": "MIDDLE"},
    {"title": "Lamb to the Slaughter by Roald Dahl (Part II)", "status": "Used", "flags": "Profanity (soft), Violence (moderate), Dark themes (moderate)", "severity": "MIDDLE"},
    {"title": "The Voice of the Enemy by Juan Villoro", "status": "Used", "flags": "Dark themes (moderate), Violence (moderate)", "severity": "MIDDLE"},
    {"title": "Harrison Bergeron by Kurt Vonnegut", "status": "Used", "flags": "Violence (moderate), Dark themes (moderate)", "severity": "MIDDLE"},
    {"title": "Interview with John Lewis (Part I)", "status": "Used", "flags": "Violence (moderate), Dark themes (moderate)", "severity": "MIDDLE"},
    {"title": "from Reading Lolita in Tehran by Azar Nafisi", "status": "Used", "flags": "Sexual indicators, Dark themes (moderate), Violence (moderate)", "severity": "MIDDLE"},
    {"title": "Reforming the World from America's Women", "status": "Used", "flags": "Sexual indicators, Violence (moderate), Dark themes (moderate)", "severity": "MIDDLE"},
    {"title": "The Feather Pillow by Horacio Quiroga", "status": "Used", "flags": "Profanity (moderate), Violence (moderate), Dark themes (moderate)", "severity": "MIDDLE"},
    {"title": "The Censors by Luisa Valenzuela", "status": "Used", "flags": "Dark themes (moderate), Violence (moderate)", "severity": "MIDDLE"},
    {"title": "There Will Come Soft Rains by Ray Bradbury", "status": "Used", "flags": "Profanity (moderate), Drug indicators, Dark themes (moderate)", "severity": "MIDDLE"},
    {"title": "By the Waters of Babylon by Stephen Vincent Benet (Part III)", "status": "Used", "flags": "Violence (moderate), Dark themes (moderate)", "severity": "MIDDLE"},
    {"title": "from An Ordinary Man by Paul Rusesabagina (Part II)", "status": "Unused", "flags": "Dark themes (moderate), Violence (moderate)", "severity": "MIDDLE"},
]

# ============================================================================
# STANDARDS MAPPING (Common Core State Standards — ELA Reading)
# ============================================================================

# Internal keys map to CCSS standard numbers (displayed as RL.X.N / RI.X.N)
STANDARD_DESCRIPTIONS = {
    "R1": "Read closely; cite textual evidence to support analysis",
    "R2": "Determine theme / central idea; summarize",
    "R3": "Analyze characters, setting, plot (RL) / individuals, events, ideas (RI)",
    "R4": "Determine meaning of words and phrases; figurative language, tone",
    "R5": "Analyze text structure and how parts relate to the whole",
    "R6": "Assess point of view or purpose; how it shapes content and style",
    "R7": "Integrate and evaluate content in diverse media and formats",
    "R8": "Evaluate argument and specific claims; assess reasoning (RI only)",
    "R9": "Compare and contrast texts; analyze how authors treat themes/topics",
    "R10": "Read and comprehend complex literary and informational texts independently",
}


def grade_band(grade):
    """Return the CCSS grade band string."""
    if grade <= 8:
        return str(grade)
    elif grade <= 10:
        return "9-10"
    else:
        return "11-12"


def ccss_label(grade, std_key, strand=None):
    """
    Format an internal standard key (R1-R10) as a CCSS label.
    strand: None = combined "RL.X.N / RI.X.N", "RL" = literature only, "RI" = informational only
    """
    num = std_key.replace("R", "")
    g = grade_band(grade)

    if strand == "RL":
        return f"RL.{g}.{num}"
    elif strand == "RI":
        return f"RI.{g}.{num}"
    elif std_key == "R8":
        return f"RI.{g}.8"
    else:
        return f"RL.{g}.{num} / RI.{g}.{num}"

# Keywords in question prompts that indicate standard alignment
STANDARD_PROMPT_KEYWORDS = {
    "R1": [r"according to", r"passage states", r"based on the (?:text|passage)", r"the (?:text|author|passage) (?:says|states|explains|describes|suggests|indicates)",
           r"support(?:s|ed)? (?:by|with)", r"evidence", r"detail", r"what does the (?:text|passage)", r"which (?:detail|statement)"],
    "R2": [r"main idea", r"theme", r"central (?:idea|message)", r"summar", r"lesson", r"moral", r"mainly about", r"best (?:describes|summarizes)",
           r"what is the (?:text|passage|story) (?:mostly |mainly )?about"],
    "R3": [r"character", r"setting", r"plot", r"event", r"happen", r"protagonist", r"conflict", r"motivat",
           r"how does .+ (?:change|develop|respond|react|feel)", r"relationship"],
    "R4": [r"meaning of", r"word .+ mean", r"figurative", r"metaphor", r"simile", r"connotation", r"the (?:word|phrase|term)",
           r"vocabulary", r"definition", r"tone", r"mood", r"imagery", r"personification", r"alliteration", r"onomatopoeia",
           r"what does .+ (?:mean|suggest|imply)", r"closest in meaning"],
    "R5": [r"(?:text|passage|story|poem) structure", r"how .+ organiz", r"(?:how|why) does the author (?:organize|structure|begin|end|arrange)",
           r"stanza", r"verse", r"rhyme scheme", r"(?:how|why) does (?:this|the) (?:section|paragraph|stanza|chapter)",
           r"purpose of (?:this|the) (?:section|paragraph|stanza)"],
    "R6": [r"point of view", r"perspective", r"narrator", r"author'?s purpose", r"first.person", r"third.person",
           r"attitude", r"bias", r"audience"],
    "R7": [r"illustration", r"diagram", r"image", r"feature", r"graphic", r"visual", r"chart", r"map", r"photograph",
           r"multimedia", r"video", r"audio"],
    "R8": [r"argument", r"claim", r"(?:evidence|reason) support", r"persuad", r"rhetor", r"counterclaim", r"logical",
           r"fallac", r"credib", r"valid"],
    "R9": [r"compare .+ (?:text|passage|story|poem|article)", r"contrast .+ (?:text|passage|story|poem|article)",
           r"both (?:texts|passages|stories|poems|authors|articles|selections)",
           r"how (?:is|are|does) .+ (?:similar|different) (?:from|to|than) .+ (?:text|passage|story)",
           r"unlike the (?:first|other|previous)", r"in common with"],
    # R10 (Range of Reading) is not directly measurable via question keywords —
    # it's about whether students read at grade-level complexity. Tracked via Lexile data instead.
    "R10": [],
}

# Unit title patterns that indicate standard coverage
STANDARD_UNIT_KEYWORDS = {
    "R1": ["key details", "citing", "textual evidence", "referencing", "pre-reading", "prereading"],
    "R2": ["theme", "main idea", "central idea", "summary", "summariz", "messages", "lessons", "morals"],
    "R3": ["character", "setting", "plot", "individuals", "events", "characterization", "interactive elements"],
    "R4": ["word meaning", "figurative language", "literary devices", "word choice", "connotation", "vocabulary",
           "determining the meaning", "language choices"],
    "R5": ["organization", "structure", "text structure", "literary techniques", "idea progression"],
    "R6": ["point of view", "narrative point", "author's purpose", "multiple accounts", "perspective",
           "dramatic irony", "cultural context"],
    "R7": ["text features", "illustrations", "multimedia", "visual", "different media", "multiple sources"],
    "R8": ["persuasive", "argument", "rhetoric", "logical fallac", "evaluate an argument"],
    "R9": ["comparing and contrasting", "fiction and non-fiction", "poetic elements", "topics, themes, forms",
           "fictional portrayals"],
    "R10": [],  # Range of Reading — measured by Lexile coverage, not unit titles
}

# ============================================================================
# LITERARY CLASSIFICATION
# ============================================================================

LITERARY_UNIT_PATTERNS = [
    r"poet(?:ry|ic)", r"stories", r"novels", r"drama", r"myths",
    r"traditional stories", r"fables", r"fairy", r"shakespeare",
    r"autobiography", r"short stories", r"julius caesar", r"twelfth night",
    r"cyrano", r"anne frank",
]

# Known literary works for grades 9-12 title matching
KNOWN_LITERARY_AUTHORS = [
    "shakespeare", "homer", "chaucer", "achebe", "fitzgerald", "orwell",
    "bradbury", "poe", "dickens", "shelley", "austen", "bronte",
    "twain", "hawthorne", "steinbeck", "lee", "golding", "salinger",
    "vonnegut", "huxley", "atwood", "morrison", "angelou", "hughes",
    "frost", "whitman", "dickinson", "wordsworth", "keats", "byron",
    "tennyson", "blake", "yeats", "eliot", "cummings", "plath",
    "sophocles", "euripides", "ovid", "virgil", "dante", "cervantes",
    "moliere", "ibsen", "chekhov", "kafka", "camus", "beckett",
    "miller", "williams", "wilde", "conrad", "joyce", "woolf",
    "faulkner", "hemingway", "dahl", "cortazar", "borges", "marquez",
    "allende", "walcott", "marvell", "owen", "gordimer", "wiesel",
    "spiegelman", "sinclair", "bierce", "jewett", "crane", "ishiguro",
    "erdrich", "jackson", "benet", "quiroga", "valenzuela", "head",
    "o'brien", "adichie", "murray", "tobin",
]

LITERARY_TITLE_PATTERNS = [
    r"\bby\b\s+[A-Z]",
    r"^from\s+",
    r"\bPart\s+[IVXLCDM\d]+",
    r"\bAct\s+[IVXLCDM\d]+",
    r"\bScene\s+[IVXLCDM\d]+",
    r"\bRetold\b",
    r"\bExcerpt\b",
]


# ============================================================================
# DATA LOADING
# ============================================================================

def load_grade_data(grade):
    """Load and return the JSON data for a grade."""
    path = GRADE_DATA_PATHS[grade]
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def get_all_assessments(data):
    """Extract all assessments with their unit titles."""
    results = []
    for unit in data.get("units", []):
        unit_title = unit.get("title", "")
        for assessment in unit.get("assessments", []):
            results.append((unit_title, assessment))
    return results


def get_assessment_title(assessment):
    """Get the display title for an assessment."""
    return (assessment.get("syllabus_metadata", {}).get("title", "")
            or assessment.get("title", ""))


def get_all_content(assessment):
    """Extract all text content (HTML) from an assessment's stimuli."""
    texts = []
    for tp in assessment.get("test_parts", []):
        for section in tp.get("sections", []):
            for item in section.get("items", []):
                stim = item.get("stimulus", {})
                if stim:
                    texts.append(stim.get("content_html", ""))
    return " ".join(texts)


def get_all_content_text(assessment):
    """Extract all plain text content from an assessment's stimuli."""
    texts = []
    for tp in assessment.get("test_parts", []):
        for section in tp.get("sections", []):
            for item in section.get("items", []):
                stim = item.get("stimulus", {})
                if stim:
                    texts.append(stim.get("content_text", ""))
    return " ".join(texts)


def get_all_questions(assessment):
    """Extract all question items from an assessment."""
    questions = []
    for tp in assessment.get("test_parts", []):
        for section in tp.get("sections", []):
            for item in section.get("items", []):
                if item.get("choices"):
                    questions.append(item)
    return questions


# ============================================================================
# TEXT CLASSIFICATION
# ============================================================================

def is_literary_unit(unit_title):
    """Check if a unit title suggests literary content."""
    title_lower = unit_title.lower()
    return any(re.search(p, title_lower) for p in LITERARY_UNIT_PATTERNS)


def has_actual_literary_content(html_content, text_content, title):
    """
    Check if the content IS actual literary text (poem, story, play)
    vs being ABOUT a literary topic (informational).
    """
    title_lower = title.lower()

    # Grades 9-12 titles often have quoted work names with genre tags
    # e.g., "The Lottery" (short story), "Dulce et Decorum Est" (poem)
    genre_match = re.search(r'["""].*?["""]\s*\((?:short story|poem|poems|novella|play|drama|myth|fable|memoir|speech|excerpt)', title, re.IGNORECASE)
    if genre_match:
        # Check if it's actually a literary genre
        genre = genre_match.group(0).lower()
        if any(g in genre for g in ["story", "poem", "novella", "play", "drama", "myth", "fable", "memoir"]):
            return True
        # Speeches and essays are informational
        if "speech" in genre or "essay" in genre:
            return False

    # Quoted title with "excerpt" = literary excerpt
    if re.search(r'["""].*?["""]\s*\(.*excerpt', title, re.IGNORECASE):
        return True

    # Title patterns that strongly indicate actual literary works
    for pattern in LITERARY_TITLE_PATTERNS:
        if re.search(pattern, title, re.IGNORECASE):
            return True

    # Check for known literary authors in title
    for author in KNOWN_LITERARY_AUTHORS:
        if author.lower() in title_lower:
            return True

    # Content-based checks
    if html_content:
        # Substantial blockquote content (poetry/literary excerpts)
        blockquotes = re.findall(r"<blockquote[^>]*>(.*?)</blockquote>", html_content, re.DOTALL)
        total_bq_len = sum(len(bq) for bq in blockquotes)
        if total_bq_len > 200:
            return True

        # Multiple line breaks within paragraphs (verse)
        br_in_p = len(re.findall(r"<br\s*/?>", html_content))
        if br_in_p > 8:
            return True

        # Play/drama indicators: character names in bold/strong followed by speech
        drama_patterns = re.findall(r"<(?:strong|b)>[A-Z][A-Z\s]+</(?:strong|b)>", html_content)
        if len(drama_patterns) > 3:
            return True

    if text_content:
        # Heavy dialogue (fiction indicator)
        dialogue = re.findall(r'["""][^"""]{5,}["""]', text_content)
        if len(dialogue) > 5:
            return True

        # Check for verse-like structure: many short lines
        lines = text_content.split("\n")
        short_lines = [l for l in lines if 5 < len(l.strip()) < 50]
        if len(short_lines) > 10 and len(short_lines) > len(lines) * 0.5:
            return True

    return False


def classify_text_type(assessment, unit_title, grade):
    """
    Classify an assessment as 'literary' or 'informational'.
    Literary = the text IS a literary work (poem, story, play, myth, etc.)
    Informational = everything else, including texts ABOUT literary works.
    """
    title = get_assessment_title(assessment)
    html_content = get_all_content(assessment)
    text_content = get_all_content_text(assessment)

    if grade <= 8:
        if is_literary_unit(unit_title):
            # Within a literary unit, check if it's actual literary content
            if has_actual_literary_content(html_content, text_content, title):
                return "literary"
            else:
                return "informational"
        else:
            # Non-literary unit: always informational
            # (even if it quotes a literary work, the assessment is informational)
            return "informational"
    else:
        # Grades 9-12: all in one unit, classify per assessment
        if has_actual_literary_content(html_content, text_content, title):
            return "literary"
        else:
            return "informational"


def classify_excerpt(assessment, unit_title, grade):
    """
    Classify if an assessment is an excerpt of a literary piece.
    Returns 'excerpt' or 'non-excerpt'.
    """
    title = get_assessment_title(assessment)
    title_lower = title.lower()

    # Title patterns indicating excerpts
    excerpt_patterns = [
        r"\bPart\s+[IVXLCDM\d]+",
        r"^from\s+",
        r"\bAct\s+[IVXLCDM\d]+",
        r"\bScene\s+[IVXLCDM\d]+",
        r"\bChapter\s+[IVXLCDM\d]+",
        r"\bExcerpt\b",
        r"\bRetold\b",
        r"\(Part\b",
        r"\bexcerpt\b",
        r"\bexcerpts?\b",
    ]

    for pattern in excerpt_patterns:
        if re.search(pattern, title, re.IGNORECASE):
            return "excerpt"

    # Check for genre tags that indicate full works: (short story), (poem), (full essay)
    # These are typically COMPLETE, not excerpts
    if re.search(r'\((?:short story|poem|poems|full essay|full text|speech)\)', title, re.IGNORECASE):
        return "non-excerpt"

    # For grades 9-12: titles with "by Author" but no Part/Act indicator
    # are likely single excerpts from longer works (novels especially)
    # But short stories and poems are usually complete
    if grade >= 9:
        # Novels and plays are usually excerpted
        if re.search(r'\(.*(?:novel|play|drama).*\)', title, re.IGNORECASE):
            return "excerpt"
        # "An Excerpt from" pattern
        if re.search(r'excerpt from', title, re.IGNORECASE):
            return "excerpt"

    # Check for known literary authors in grades 3-8 (suggests excerpt)
    if grade <= 8:
        for author in KNOWN_LITERARY_AUTHORS:
            if author.lower() in title_lower:
                return "excerpt"

    return "non-excerpt"


def classify_originality(assessment, unit_title, grade):
    """
    For assessments in literary units (grades 3-8), determine whether the
    primary text is the ACTUAL original literary work or a curriculum-written
    synopsis/retelling/educational article.

    The key structural pattern in this curriculum:
    - Synopsis articles: ALL stimuli have educational H2 subheadings and explanatory framing
    - Genuine excerpt assessments: title contains "Excerpt"/"Complete Story" etc., and
      stimulus 2 contains actual literary text (often under heading "The Excerpt")

    Returns:
      'original'  — the assessment presents actual original text as the primary reading
      'synopsis_with_quotes' — curriculum-written article with embedded original excerpts
      'synopsis'  — entirely curriculum-written synopsis/retelling/educational article
      None        — not in a literary unit (not applicable)
    """
    if grade > 8:
        return None  # Grades 9-12 generally present actual texts

    if not is_literary_unit(unit_title):
        return None  # Only analyze literary-unit assessments

    title = get_assessment_title(assessment)

    # Collect all stimuli across all sections
    stimuli = []
    for tp in assessment.get("test_parts", []):
        for section in tp.get("sections", []):
            for item in section.get("items", []):
                stim = item.get("stimulus")
                if stim and stim.get("content_text", "").strip():
                    stimuli.append(stim)

    if not stimuli:
        return "synopsis"

    # Title indicators of actual excerpt/original content
    is_excerpt_title = bool(re.search(
        r"\b(?:Excerpt|An Excerpt from|Complete (?:Short )?Story|Complete Poem|The Scene:)\b",
        title, re.IGNORECASE
    ))

    # Check stimulus 2 (the most important one) for "The Excerpt" pattern
    has_excerpt_stimulus = False
    if len(stimuli) >= 2:
        stim2_text = stimuli[1].get("content_text", "")
        stim2_html = stimuli[1].get("content_html", "")
        # Look for "The Excerpt" or "The Scene" heading
        if re.search(r"(?:The Excerpt|The Scene|The Passage|The Speech|The Poem)", stim2_text[:100]):
            has_excerpt_stimulus = True
        # Check for source attribution in the stimulus
        if re.search(r"(?:Source:\s*[A-Z]|— [A-Z][a-z]+ [A-Z]|Project Gutenberg)", stim2_text):
            has_excerpt_stimulus = True

    # If title says excerpt AND stimulus 2 has actual text → original
    if is_excerpt_title and has_excerpt_stimulus:
        return "original"
    if is_excerpt_title:
        # Title says excerpt but check content to confirm
        if len(stimuli) >= 2:
            stim2_len = len(stimuli[1].get("content_text", ""))
            stim2_html = stimuli[1].get("content_html", "")
            # Long stimulus 2 with dramatic dialogue or verse = likely original
            drama_lines = len(re.findall(r"<(?:strong|b)>[A-Z][A-Z\s]+</(?:strong|b)>", stim2_html))
            br_count = len(re.findall(r"<br\s*/?>", stim2_html))
            if drama_lines > 3 or (br_count > 6 and stim2_len > 500):
                return "original"
            if stim2_len > 2000:
                return "original"
        return "original"  # Trust the title

    # For non-excerpt titles: these are curriculum-written articles.
    # Check if they embed any original quotes (blockquotes) within the synopsis.
    total_bq_len = 0
    for stim in stimuli:
        html = stim.get("content_html", "")
        blockquotes = re.findall(r"<blockquote[^>]*>(.*?)</blockquote>", html, re.DOTALL)
        total_bq_len += sum(len(bq) for bq in blockquotes)

    # Even with embedded quotes, the primary reading is the curriculum-written article.
    # Classify based on how much original text is embedded.
    if total_bq_len > 300:
        return "synopsis_with_quotes"

    return "synopsis"


# ============================================================================
# ANSWER DISTRIBUTION ANALYSIS
# ============================================================================

def analyze_answer_distribution(assessments):
    """
    Count correct answer positions across all questions.
    Returns dict of {position: count} where position is 0-3 (A-D).
    """
    position_counts = Counter()
    total = 0

    for _, assessment in assessments:
        questions = get_all_questions(assessment)
        for q in questions:
            choices = q.get("choices", [])
            for i, choice in enumerate(choices):
                if choice.get("is_correct"):
                    position_counts[i] += 1
                    total += 1
                    break

    return position_counts, total


# ============================================================================
# STANDARDS COVERAGE ANALYSIS
# ============================================================================

def analyze_standards_coverage(assessments, grade):
    """
    Analyze which standards (R1-R10) are covered by questions and units,
    separated into RL (Reading Literature) and RI (Reading Informational).

    Returns dict of:
      {standard: {
          "rl_questions": n, "ri_questions": n,
          "rl_assessments": n, "ri_assessments": n,
          "total_questions": n, "total_assessments": n,
          "total_rl_questions": n, "total_ri_questions": n,
          "unit_titles": set
      }}
    """
    coverage = {}
    for std in STANDARD_DESCRIPTIONS:
        coverage[std] = {
            "rl_questions": 0,
            "ri_questions": 0,
            "rl_assessments": 0,
            "ri_assessments": 0,
            "total_questions": 0,
            "total_assessments": 0,
            "total_rl_questions": 0,
            "total_ri_questions": 0,
            "unit_titles": set(),
        }

    for unit_title, assessment in assessments:
        questions = get_all_questions(assessment)
        text_type = classify_text_type(assessment, unit_title, grade)
        is_literary = (text_type == "literary")
        assessment_standards = set()

        # Check unit title against standard keywords
        unit_lower = unit_title.lower()
        for std, keywords in STANDARD_UNIT_KEYWORDS.items():
            if any(kw in unit_lower for kw in keywords):
                assessment_standards.add(std)
                coverage[std]["unit_titles"].add(unit_title)

        # Check each question prompt against standard keywords
        for q in questions:
            prompt = (q.get("prompt", "") or "").lower()
            for std, patterns in STANDARD_PROMPT_KEYWORDS.items():
                if any(re.search(p, prompt) for p in patterns):
                    if is_literary:
                        coverage[std]["rl_questions"] += 1
                    else:
                        coverage[std]["ri_questions"] += 1
                    assessment_standards.add(std)

        for std in assessment_standards:
            if is_literary:
                coverage[std]["rl_assessments"] += 1
            else:
                coverage[std]["ri_assessments"] += 1

        for std in coverage:
            coverage[std]["total_questions"] += len(questions)
            coverage[std]["total_assessments"] += 1
            if is_literary:
                coverage[std]["total_rl_questions"] += len(questions)
            else:
                coverage[std]["total_ri_questions"] += len(questions)

    return coverage


# ============================================================================
# FLAGGED ITEMS CHECK
# ============================================================================

def extract_core_title(title):
    """
    Extract the core work title from various formatting styles.
    Handles: "The Lottery" (short story), The Lottery by Shirley Jackson (Part II),
    from The Lottery, 'The Most Dangerous Game', etc.
    """
    t = title.strip()

    # Extract title from any type of quotes: "Title", "Title", 'Title'
    quoted = re.search(r'[\u201c\u201d"""\u2018\u2019\u0027]([^\u201c\u201d"""\u2018\u2019\u0027]{3,})[\u201c\u201d"""\u2018\u2019\u0027]', t)
    if quoted:
        t = quoted.group(1)

    # Remove "from " prefix
    t = re.sub(r"^from\s+", "", t, flags=re.IGNORECASE)

    # Remove "An Excerpt from" or "Excerpt from"
    t = re.sub(r"^(?:An\s+)?Excerpt\s+from\s+", "", t, flags=re.IGNORECASE)

    # Remove "by Author Name" suffix
    t = re.sub(r"\s+by\s+[A-Z][a-zA-Z\s.,'\u2019-]+$", "", t)

    # Remove part/act/scene indicators
    t = re.sub(r"\s*[\(,]\s*Part\s+[IVXLCDM\d]+\s*\)?", "", t, flags=re.IGNORECASE)
    t = re.sub(r"\s*,?\s*Act\s+[IVXLCDM\d]+.*$", "", t, flags=re.IGNORECASE)
    t = re.sub(r"\s*,?\s*Scene\s+[IVXLCDM\d]+.*$", "", t, flags=re.IGNORECASE)
    t = re.sub(r"\s*,?\s*Scenes?\s+\d+.*$", "", t, flags=re.IGNORECASE)
    t = re.sub(r"\s*,?\s*Chapter\s+[IVXLCDM\d]+.*$", "", t, flags=re.IGNORECASE)

    # Remove genre tags: (short story), (poem), (essay), (excerpt), (full essay, ...)
    t = re.sub(r"\s*\([^)]*(?:story|poem|essay|excerpt|speech|novel|play|drama|memoir|retelling|meditation|paired|revisit)[^)]*\)\s*$", "", t, flags=re.IGNORECASE)

    # Strip ALL quotation marks (straight and curly, single and double)
    t = re.sub(r'[\u201c\u201d\u201e\u201f\u2018\u2019\u201a\u201b""\'`]', '', t)

    # Remove trailing punctuation and whitespace
    t = t.strip().rstrip(".,;:!?—–-")
    t = re.sub(r"\s+", " ", t).strip()

    return t.lower()


def normalize_title(title):
    """Normalize a title for fuzzy matching."""
    t = title.lower().strip()
    t = re.sub(r'[\u201c\u201d\u201e\u201f\u2018\u2019\u201a\u201b""\'`\u0027]', "", t)
    t = re.sub(r"\s+", " ", t)
    t = t.rstrip(".,;:!?")
    return t


def find_flagged_items(assessments, grade):
    """
    Check if any previously flagged items are present in the current data.
    Uses core title extraction for matching.
    """
    if grade < 5:
        return []

    # Build lookup of current assessment titles with core title extraction
    current_titles = {}  # core_title -> original_title
    current_normalized = {}  # normalized -> original
    for unit_title, assessment in assessments:
        title = get_assessment_title(assessment)
        core = extract_core_title(title)
        if core:
            current_titles[core] = title
        current_normalized[normalize_title(title)] = title

    found = []
    for flagged in FLAGGED_ITEMS:
        flagged_core = extract_core_title(flagged["title"])
        flagged_norm = normalize_title(flagged["title"])

        matched_title = None
        match_type = None

        # Core title match (most reliable)
        if flagged_core in current_titles:
            matched_title = current_titles[flagged_core]
            match_type = "core"

        # Exact normalized match
        if not matched_title and flagged_norm in current_normalized:
            matched_title = current_normalized[flagged_norm]
            match_type = "exact"

        # Substring match on core titles
        if not matched_title and len(flagged_core) > 8:
            for curr_core, curr_orig in current_titles.items():
                if len(curr_core) > 8 and (flagged_core in curr_core or curr_core in flagged_core):
                    matched_title = curr_orig
                    match_type = "partial"
                    break

        if matched_title:
            found.append({
                "flagged_title": flagged["title"],
                "current_title": matched_title,
                "status": flagged["status"],
                "flags": flagged["flags"],
                "severity": flagged["severity"],
                "match_type": match_type,
            })

    # Deduplicate by flagged_title
    seen = set()
    deduped = []
    for item in found:
        if item["flagged_title"] not in seen:
            seen.add(item["flagged_title"])
            deduped.append(item)

    return deduped


# ============================================================================
# CURRICULUM PLAN PARSING (High School XLSX)
# ============================================================================

CURRICULUM_PLAN_PATH = "/Users/alexandra/Claude Code/Alpha Read Analysis/high_school_reading_curriculum.xlsx"

# Planned Lexile ranges from curriculum plan
PLANNED_LEXILE = {
    9: {"range": "1050L–1260L", "midpoint": 1155, "early": "1050L–1120L", "mid": "1120L–1190L", "late": "1190L–1260L"},
    10: {"range": "1080L–1335L", "midpoint": 1205, "early": "1080L–1165L", "mid": "1165L–1250L", "late": "1250L–1335L"},
    11: {"range": "1185L–1385L", "midpoint": 1285, "early": "1185L–1250L", "mid": "1250L–1320L", "late": "1320L–1385L"},
    12: {"range": "1250L–1420L", "midpoint": 1335, "early": "1250L–1305L", "mid": "1305L–1365L", "late": "1365L–1420L"},
}

# Planned article counts from overview sheet
PLANNED_COUNTS = {
    9: {"articles": 130, "units": 7, "literary": 71, "informational": 59},
    10: {"articles": 135, "units": 6, "literary": 66, "informational": 69},
    11: {"articles": 135, "units": 6, "literary": 69, "informational": 66},
    12: {"articles": 130, "units": 6, "literary": 80, "informational": 50},
}

# Planned word count ranges
PLANNED_WORD_COUNTS = {
    9: "1,700–2,100", 10: "1,800–2,200", 11: "1,900–2,350", 12: "2,000–2,500",
}


def parse_curriculum_plan():
    """Parse the high school curriculum XLSX to extract planned texts per grade."""
    if not HAS_OPENPYXL or not os.path.exists(CURRICULUM_PLAN_PATH):
        return {}

    wb = load_workbook(CURRICULUM_PLAN_PATH, data_only=True, read_only=True)
    plan = {}

    for grade in [9, 10, 11, 12]:
        sheet_name = f"Grade {grade} Curriculum"
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        texts = []
        current_unit = ""
        custom_count = 0
        genres = Counter()

        for row in rows:
            # Skip header / empty rows
            if not row or len(row) < 6:
                continue

            # Detect unit headers (column B has unit name, column C is empty)
            col_b = str(row[1] or "").strip()
            col_c = str(row[2] or "").strip()
            col_d = str(row[3] or "").strip()  # Text Title
            col_e = str(row[4] or "").strip()  # Author
            col_f = str(row[5] or "").strip()  # Text Type
            col_g = str(row[6] or "").strip() if len(row) > 6 else ""  # Complexity
            col_h = str(row[7] or "").strip() if len(row) > 7 else ""  # Notes

            # Detect unit row: starts with "Unit" in col_b
            if col_b.startswith("Unit ") and not col_d:
                current_unit = col_b
                continue

            # Skip meta rows
            if col_b in ("Unit", "TOTAL", "") and not col_d:
                continue
            if col_b.startswith("Grade ") or col_b.startswith("Lexile:"):
                continue
            if col_b.startswith("GRADE "):
                continue

            # Text row: col_d has title
            if col_d and col_d not in ("Text Title", "NaN", "None", "nan"):
                is_custom = "Custom Informational Text" in col_e or "CUSTOM TEXT NEEDED" in col_h
                text_type = col_f if col_f and col_f not in ("None", "nan", "NaN") else ""
                is_literary = "Literary" in text_type
                is_informational = "Informational" in text_type or is_custom

                if is_custom:
                    custom_count += 1

                # Extract genre
                genre = ""
                if "Poetry" in text_type:
                    genre = "Poetry"
                elif "Drama" in text_type:
                    genre = "Drama"
                elif "Fiction" in text_type:
                    genre = "Fiction"
                elif "Memoir" in text_type or "Autobiography" in text_type:
                    genre = "Memoir"
                elif "Speech" in text_type:
                    genre = "Speech"
                elif "Essay" in text_type:
                    genre = "Essay"
                elif "Expository" in text_type:
                    genre = "Expository"
                elif "History" in text_type:
                    genre = "History"
                elif "Science" in text_type:
                    genre = "Science"
                elif "Philosophy" in text_type:
                    genre = "Philosophy"
                else:
                    genre = "Other"

                genres[genre] += 1

                texts.append({
                    "title": col_d,
                    "author": col_e,
                    "text_type": text_type,
                    "is_literary": is_literary,
                    "is_informational": is_informational,
                    "is_custom": is_custom,
                    "complexity": col_g,
                    "unit": current_unit,
                    "notes": col_h[:200] if col_h else "",
                })

        # Extract planned units
        planned_units = []
        seen_units = set()
        for t in texts:
            if t["unit"] and t["unit"] not in seen_units:
                seen_units.add(t["unit"])
                planned_units.append(t["unit"])

        plan[grade] = {
            "texts": texts,
            "total": len(texts),
            "custom_count": custom_count,
            "literary_count": sum(1 for t in texts if t["is_literary"]),
            "informational_count": sum(1 for t in texts if t["is_informational"]),
            "genres": dict(genres),
            "units": planned_units,
        }

    wb.close()
    return plan


def cross_reference_plan(plan_texts, assessments, grade):
    """Cross-reference planned curriculum texts against actual QTI assessments."""
    if not plan_texts:
        return {}

    # Build actual title lookup
    actual_cores = {}
    actual_normalized = {}
    for unit_title, assessment in assessments:
        title = get_assessment_title(assessment)
        core = extract_core_title(title)
        if core:
            actual_cores[core] = title
        actual_normalized[normalize_title(title)] = title

    found_in_qti = []
    missing_from_qti = []
    custom_not_yet_written = []

    for planned in plan_texts:
        p_title = planned["title"]
        p_core = extract_core_title(p_title)
        p_norm = normalize_title(p_title)

        matched = None
        match_type = None

        # Core title match
        if p_core and p_core in actual_cores:
            matched = actual_cores[p_core]
            match_type = "core"

        # Normalized match
        if not matched and p_norm in actual_normalized:
            matched = actual_normalized[p_norm]
            match_type = "exact"

        # Substring match
        if not matched and p_core and len(p_core) > 10:
            for a_core, a_orig in actual_cores.items():
                if len(a_core) > 10 and (p_core in a_core or a_core in p_core):
                    matched = a_orig
                    match_type = "partial"
                    break

        if matched:
            found_in_qti.append({
                "planned_title": p_title,
                "actual_title": matched,
                "match_type": match_type,
                "is_custom": planned["is_custom"],
                "text_type": planned["text_type"],
            })
        elif planned["is_custom"]:
            custom_not_yet_written.append(planned)
        else:
            missing_from_qti.append(planned)

    # Find QTI assessments not in the plan
    found_plan_cores = set()
    for f in found_in_qti:
        found_plan_cores.add(extract_core_title(f["actual_title"]))

    extra_in_qti = []
    for unit_title, assessment in assessments:
        title = get_assessment_title(assessment)
        core = extract_core_title(title)
        if core and core not in found_plan_cores:
            extra_in_qti.append(title)

    return {
        "found_in_qti": found_in_qti,
        "missing_from_qti": missing_from_qti,
        "custom_not_yet_written": custom_not_yet_written,
        "extra_in_qti": extra_in_qti,
        "match_rate": len(found_in_qti) / len(plan_texts) * 100 if plan_texts else 0,
    }


# ============================================================================
# STRENGTHS & WEAKNESSES
# ============================================================================

def analyze_grade(grade, data):
    """Run all analyses for a single grade. Returns a metrics dict."""
    assessments = get_all_assessments(data)
    total_articles = len(assessments)

    # Text type classification
    literary_count = 0
    informational_count = 0
    excerpt_count = 0
    non_excerpt_count = 0
    literary_titles = []
    informational_titles = []
    excerpt_titles = []

    # Originality classification (grades 3-8 literary units only)
    originality_counts = Counter()  # 'original', 'synopsis_with_quotes', 'synopsis'
    originality_details = {"original": [], "synopsis_with_quotes": [], "synopsis": []}
    total_literary_unit_assessments = 0

    for unit_title, assessment in assessments:
        title = get_assessment_title(assessment)
        text_type = classify_text_type(assessment, unit_title, grade)
        excerpt_type = classify_excerpt(assessment, unit_title, grade)

        if text_type == "literary":
            literary_count += 1
            literary_titles.append(title)
        else:
            informational_count += 1
            informational_titles.append(title)

        if excerpt_type == "excerpt":
            excerpt_count += 1
            excerpt_titles.append(title)
        else:
            non_excerpt_count += 1

        # Originality check for literary units
        originality = classify_originality(assessment, unit_title, grade)
        if originality is not None:
            total_literary_unit_assessments += 1
            originality_counts[originality] += 1
            originality_details[originality].append(title)

    # Answer distribution
    answer_dist, total_questions = analyze_answer_distribution(assessments)

    # Standards coverage
    standards = analyze_standards_coverage(assessments, grade)

    # Flagged items
    flagged = find_flagged_items(assessments, grade)

    # Unit info
    units = data.get("units", [])
    unit_titles = [u["title"] for u in units]

    # Question type analysis
    question_types = Counter()
    section_types = Counter()
    total_bold_words = 0
    total_stimuli = 0
    has_glossary = False
    has_graphic_organizer = False

    for unit_title, assessment in assessments:
        for tp in assessment.get("test_parts", []):
            for section in tp.get("sections", []):
                sec_title = section.get("title", "")
                if "guiding" in sec_title.lower():
                    section_types["guiding"] += 1
                elif "quiz" in sec_title.lower():
                    section_types["quiz"] += 1
                else:
                    section_types["other"] += 1

                for item in section.get("items", []):
                    question_types[item.get("interaction_type", "choice")] += 1
                    stim = item.get("stimulus", {})
                    if stim:
                        total_stimuli += 1
                        html = stim.get("content_html", "")
                        bold_count = len(re.findall(r"<strong>", html))
                        total_bold_words += bold_count
                        if "glossary" in html.lower() or "vocabulary list" in html.lower():
                            has_glossary = True
                        if "graphic organizer" in html.lower():
                            has_graphic_organizer = True

    # Lexile data
    lexile_values = []
    for _, assessment in assessments:
        lex = assessment.get("metadata", {}).get("lexileLevel", "")
        if lex and lex != "0":
            try:
                lexile_values.append(int(lex))
            except ValueError:
                pass

    # Compute strengths and weaknesses
    literary_pct = (literary_count / total_articles * 100) if total_articles else 0
    informational_pct = (informational_count / total_articles * 100) if total_articles else 0
    excerpt_pct = (excerpt_count / total_articles * 100) if total_articles else 0

    strengths = []
    weaknesses = []

    # --- STRENGTHS ---
    if total_articles > 80:
        strengths.append(f"Large volume of reading material ({total_articles} articles)")

    if len(units) > 10:
        strengths.append(f"Well-organized thematic unit structure ({len(units)} units covering diverse topics)")

    if section_types.get("guiding", 0) > 0:
        strengths.append("Scaffolded reading with guiding questions before synthesis quiz")

    # Check for feedback on answers
    has_feedback = False
    for _, assessment in assessments[:5]:
        qs = get_all_questions(assessment)
        for q in qs:
            for c in q.get("choices", []):
                if c.get("feedback"):
                    has_feedback = True
                    break
    if has_feedback:
        strengths.append("Detailed feedback provided for every answer choice (correct and incorrect)")

    if lexile_values:
        avg_lexile = sum(lexile_values) / len(lexile_values)
        strengths.append(f"Lexile-leveled texts (avg Lexile: {avg_lexile:.0f})")

    if informational_pct > 40 and grade <= 8:
        strengths.append("Strong cross-curricular informational content (history, science, arts)")

    if literary_pct > 0:
        strengths.append(f"Includes literary texts ({literary_pct:.0f}% of content)")

    # Check for diverse topics
    topic_areas = set()
    for ut in unit_titles:
        ut_lower = ut.lower()
        if any(w in ut_lower for w in ["science", "human body", "chemistry", "physics", "ecology", "astronomy"]):
            topic_areas.add("Science")
        if any(w in ut_lower for w in ["history", "revolution", "war", "civilization", "colonial"]):
            topic_areas.add("History")
        if any(w in ut_lower for w in ["art", "painting", "sculpture", "architecture"]):
            topic_areas.add("Visual Arts")
        if any(w in ut_lower for w in ["music", "composer", "orchestra", "song"]):
            topic_areas.add("Music")
        if any(w in ut_lower for w in ["geography", "canada", "africa", "europe", "america"]):
            topic_areas.add("Geography")
        if any(w in ut_lower for w in ["poetry", "stories", "novels", "drama"]):
            topic_areas.add("Literature")
    if len(topic_areas) >= 4:
        strengths.append(f"Interdisciplinary coverage spans {len(topic_areas)} subject areas: {', '.join(sorted(topic_areas))}")

    # --- WEAKNESSES ---

    # Answer distribution bias
    if total_questions > 0:
        a_pct = answer_dist.get(0, 0) / total_questions * 100
        b_pct = answer_dist.get(1, 0) / total_questions * 100
        c_pct = answer_dist.get(2, 0) / total_questions * 100
        d_pct = answer_dist.get(3, 0) / total_questions * 100
        ab_combined = a_pct + b_pct
        if ab_combined > 75:
            weaknesses.append(
                f"SEVERE answer key bias: {ab_combined:.0f}% of correct answers are A or B "
                f"(A={a_pct:.0f}%, B={b_pct:.0f}%, C={c_pct:.0f}%, D={d_pct:.0f}%). "
                f"Students can achieve above-chance scores by always selecting A or B. "
                f"This undermines assessment validity and teaches test-taking shortcuts."
            )

    # Literary/informational balance
    if grade <= 5:
        # CCSS recommends ~50/50 literary/informational for elementary
        if literary_pct < 30:
            weaknesses.append(
                f"Literary texts comprise only {literary_pct:.0f}% of content. "
                f"Common Core recommends approximately 50% literary / 50% informational for grades 3-5. "
                f"Students have limited exposure to actual poetry, fiction, and drama."
            )
    elif grade <= 8:
        if literary_pct < 35:
            weaknesses.append(
                f"Literary texts comprise only {literary_pct:.0f}% of content. "
                f"Middle school should maintain substantial literary text engagement for "
                f"developing interpretive and analytical skills."
            )

    # High excerpt percentage
    if excerpt_pct > 60:
        weaknesses.append(
            f"{excerpt_pct:.0f}% of texts are excerpts of longer works. Students rarely encounter "
            f"complete texts, limiting their ability to analyze full narrative arcs, "
            f"character development, and thematic resolution."
        )

    # Missing Lexile data
    if not lexile_values and grade >= 9:
        weaknesses.append(
            "No Lexile level data for any text. Without readability metrics, it's impossible to "
            "verify grade-level appropriateness or differentiate texts by complexity."
        )

    # Single unit structure (grades 9-12)
    if len(units) <= 1:
        weaknesses.append(
            "All content placed in a single undifferentiated unit ('Reading Exercises'). "
            "No thematic organization, no progression from simpler to complex texts, "
            "no genre-based grouping. This makes curriculum pacing and targeted instruction impossible."
        )

    # Standards coverage gaps — check RL and RI separately
    uncovered_rl = []
    uncovered_ri = []
    weakly_covered_rl = []
    weakly_covered_ri = []

    for std in ["R1", "R2", "R3", "R4", "R5", "R6", "R7", "R9"]:
        s = standards[std]
        if s["rl_questions"] == 0 and s["rl_assessments"] == 0:
            uncovered_rl.append(std)
        elif s["rl_questions"] < 5:
            weakly_covered_rl.append(std)

    for std in ["R1", "R2", "R3", "R4", "R5", "R6", "R7", "R8", "R9"]:
        s = standards[std]
        if s["ri_questions"] == 0 and s["ri_assessments"] == 0:
            uncovered_ri.append(std)
        elif s["ri_questions"] < 5:
            weakly_covered_ri.append(std)

    if uncovered_rl:
        std_names = ", ".join(f"{ccss_label(grade, s, 'RL')}" for s in uncovered_rl)
        weaknesses.append(f"RL (Literature) standards with zero coverage: {std_names}")
    if uncovered_ri:
        std_names = ", ".join(f"{ccss_label(grade, s, 'RI')}" for s in uncovered_ri)
        weaknesses.append(f"RI (Informational) standards with zero coverage: {std_names}")
    if weakly_covered_rl:
        std_names = ", ".join(f"{ccss_label(grade, s, 'RL')} ({standards[s]['rl_questions']}q)" for s in weakly_covered_rl)
        weaknesses.append(f"RL standards with minimal coverage (<5 questions): {std_names}")
    if weakly_covered_ri:
        std_names = ", ".join(f"{ccss_label(grade, s, 'RI')} ({standards[s]['ri_questions']}q)" for s in weakly_covered_ri)
        weaknesses.append(f"RI standards with minimal coverage (<5 questions): {std_names}")

    # R10 (Range of Reading) — check Lexile
    if not lexile_values:
        weaknesses.append(
            f"{ccss_label(grade, 'R10')}: No Lexile data to verify grade-level text complexity. "
            f"Cannot assess whether students are reading at appropriate complexity levels."
        )

    # Flagged items
    if flagged:
        high_severity = [f for f in flagged if f["severity"] == "HIGH"]
        previously_removed = [f for f in flagged if f["status"] == "Unused"]
        if previously_removed:
            weaknesses.append(
                f"{len(previously_removed)} previously removed/flagged article(s) appear to have been "
                f"re-added to the curriculum despite being flagged for inappropriate content."
            )
        if high_severity:
            weaknesses.append(
                f"{len(high_severity)} article(s) with HIGH-severity content flags "
                f"(severe violence, dark themes, sexual content) are currently present."
            )

    # Originality: synopsis vs original text (grades 3-8)
    if grade <= 8 and total_literary_unit_assessments > 0:
        synopsis_count = originality_counts.get("synopsis", 0)
        synopsis_pct = synopsis_count / total_literary_unit_assessments * 100
        original_count = originality_counts.get("original", 0)
        hybrid_count = originality_counts.get("synopsis_with_quotes", 0)
        if synopsis_pct > 50:
            weaknesses.append(
                f"CRITICAL: {synopsis_count} of {total_literary_unit_assessments} literary-unit assessments "
                f"({synopsis_pct:.0f}%) are curriculum-written synopses, retellings, or educational articles "
                f"— NOT the original literary texts. Only {original_count} present actual original text "
                f"and {hybrid_count} embed brief original quotes within synopsis framing. "
                f"Students overwhelmingly read ABOUT literature rather than reading literature itself."
            )

    return {
        "grade": grade,
        "total_articles": total_articles,
        "total_questions": total_questions,
        "num_units": len(units),
        "unit_titles": unit_titles,
        "literary_count": literary_count,
        "informational_count": informational_count,
        "literary_pct": literary_pct,
        "informational_pct": informational_pct,
        "excerpt_count": excerpt_count,
        "non_excerpt_count": non_excerpt_count,
        "excerpt_pct": excerpt_pct,
        "non_excerpt_pct": 100 - excerpt_pct,
        "answer_distribution": answer_dist,
        "answer_total": total_questions,
        "standards": standards,
        "flagged_items": flagged,
        "strengths": strengths,
        "weaknesses": weaknesses,
        "question_types": dict(question_types),
        "section_types": dict(section_types),
        "lexile_values": lexile_values,
        "literary_titles": literary_titles,
        "informational_titles": informational_titles,
        "excerpt_titles": excerpt_titles,
        "originality_counts": dict(originality_counts),
        "originality_details": originality_details,
        "total_literary_unit_assessments": total_literary_unit_assessments,
    }


# ============================================================================
# CONTEXTUALIZATION ANALYSIS
# ============================================================================

# Texts that need contextual introductions, organized by reason category.
# Each entry: (title_fragment, reason, category)
# Categories: "satire", "ancient", "mid-excerpt", "primary-source", "philosophical"
TEXTS_NEEDING_CONTEXT = [
    # --- Satire / Irony that will be misread ---
    ("Modest Proposal", "Satirical essay proposing eating children; students may take it literally without knowing the Irish famine context and Swift's satirical intent", "satire"),
    ("White Man's Burden", "Imperialist ideology poem; without framing, students may read it as sincere praise of colonialism rather than understanding it as a text to critically examine", "satire"),
    ("Gulliver's Travels", "Swift's darkest satire — rational horses vs. bestial humans is absurd nonsense without knowing Swift's satirical project against human nature", "satire"),

    # --- Ancient / Archaic texts with major cultural distance ---
    ("Epic of Gilgamesh", "~4,000-year-old Mesopotamian epic; unfamiliar cosmology, names, geography; students need to know this is the oldest surviving literature", "ancient"),
    ("Iliad", "Ancient Greek epic; students need to know the Trojan War, who Achilles and Hector/Priam are, and the Greek concept of honor (kleos)", "ancient"),
    ("Ramayana", "Hindu epic; concepts of dharma, the caste system, and the epic's cultural significance in South Asia are essential for comprehension", "ancient"),
    ("Beowulf", "Anglo-Saxon warrior culture, comitatus, mead-halls — all unfamiliar; even in Heaney's translation, the style is dense and archaic", "ancient"),
    ("Antigone", "Students must know the Oedipus backstory — Antigone's brothers killed each other, Creon's edict — or the play's central conflict is unintelligible", "ancient"),
    ("Allegory of the Cave", "Philosophical allegory from 380 BC; without knowing Plato's epistemology framework, students will just see a story about people in a cave", "ancient"),
    ("In a Bamboo Grove", "12th-century Japanese setting; the Rashomon unreliable-narrator structure needs framing for students to grasp the experiment", "ancient"),
    ("Rashomon", "12th-century Japanese setting; the Rashomon unreliable-narrator structure needs framing for students to grasp the experiment", "ancient"),
    ("Prince, Chapters", "Renaissance political philosophy; without the Italian city-state context and that this was practical advice to a real ruler, students miss the point", "ancient"),
    ("Leviathan", "Dense 17th-century political philosophy; the 'state of nature' concept and social contract theory need pre-teaching", "ancient"),
    ("Don Quixote", "The first modern novel; students need to know Quixote has been driven mad by chivalric romances — the comedy depends on this irony", "ancient"),
    ("Paradise Lost", "Milton's blank verse is extremely dense; students need the biblical Fall framework and the radical idea that Satan is the protagonist", "ancient"),
    ("Rime of the Ancient Mariner", "Archaic diction, supernatural framework, and Romantic-era ballad conventions all need framing before reading", "ancient"),
    ("Utopia", "The satirical intent (is More serious or ironic?) is famously ambiguous; students need the 16th-century English context to engage", "ancient"),
    ("Excerpt from We", "1924 Soviet-era Russian dystopia; students need to know it was banned in the USSR, influenced Orwell, and pioneered the genre", "ancient"),
    ("Season of Migration", "Sudanese postcolonial novel deliberately reverses Heart of Darkness; without knowing Conrad's work, the inversion is lost", "ancient"),
    ("Metamorphosis", "Kafkaesque alienation as a literary concept; without framing, the insect transformation reads as horror rather than existentialist parable", "ancient"),
    ("Metamorphoses", "Ovid's Roman mythology collection (8 AD); students need to know the cultural context and that this is a source for Romeo and Juliet", "ancient"),
    ("Apology of Socrates", "Students need to know Socrates was on trial for his life and what the charges of impiety and corrupting youth meant in Athens", "philosophical"),
    ("Myth of Sisyphus", "Absurdist philosophy; without the framework, students will be confused by the claim that we should 'imagine Sisyphus happy'", "philosophical"),
    ("Waiting for Godot", "Theatre of the Absurd conventions; without knowing this is intentionally plotless, students will think the play is broken", "philosophical"),
    ("Stranger", "Meursault's detachment will read as bad writing rather than existentialist philosophy without a Camus/absurdism introduction", "philosophical"),
    ("Mrs. Dalloway", "Stream of consciousness with no traditional plot markers; students need to know about Woolf's technique and the shell-shock parallel narrative (Septimus)", "philosophical"),

    # --- Mid-work excerpts (dropped into middle of longer works) ---
    ("Odyssey, Books 21", "Drops into the climactic return scene; students need to know who Odysseus is, the Trojan War, the suitors occupying his palace, and Penelope's test", "mid-excerpt"),
    ("Odyssey, Books 9", "Students need the Trojan War context, Odysseus's identity, and that these are stories he tells while detained by the Phaeacians", "mid-excerpt"),
    ("Mockingbird, Chapters 17", "The trial scene requires knowing Tom Robinson, Atticus's role, the Ewell family, and Maycomb's racial dynamics", "mid-excerpt"),
    ("Lesson Before Dying, Chapters 12", "Jefferson's journal entries are meaningless without knowing his conviction, Grant's visits, and the devastating 'hog' speech", "mid-excerpt"),
    ("1984, Part 2, Chapter 9", "Goldstein's book and the final Room 101 scene; without Part 1, students won't understand Winston, Big Brother, or what's at stake", "mid-excerpt"),
    ("1984, Part 2 (Chapters 1", "Winston and Julia's rebellion requires knowing the world of Oceania, the Party, and Winston's growing dissent from Part 1", "mid-excerpt"),
    ("1984, Part 3", "Room 101 and the conclusion; the total destruction of Winston's identity requires understanding the full arc of his resistance", "mid-excerpt"),
    ("Julius Caesar, Act III", "The assassination and funeral speeches require knowing the conspirators, Caesar's rise, and Roman Republic political dynamics", "mid-excerpt"),
    ("Life of Pi, Chapters 37", "The Pacific survival narrative is incomprehensible without knowing Pi's background, the zoo, and the shipwreck", "mid-excerpt"),
    ("Life of Pi, Chapters 96", "The 'two stories' climax — 'which is the better story?' — requires the entire preceding narrative to have impact", "mid-excerpt"),
    ("Heart of Darkness, Chapters 1", "Conrad's frame narrative and delayed entrance of Kurtz; students need the colonial Congo context and Marlow's role", "mid-excerpt"),
    ("Frankenstein, Chapters 11", "The Creature's narrative requires knowing Victor's creation and abandonment from earlier chapters", "mid-excerpt"),
    ("Frankenstein, Chapters 20", "The consequences and tragedy require knowing Victor's promise to create a mate and his decision to destroy it", "mid-excerpt"),
    ("Macbeth, Act II", "The murder and its consequences require Act I's prophecy, Lady Macbeth's persuasion, and Macbeth's ambition", "mid-excerpt"),
    ("Macbeth, Act V", "'Tomorrow and tomorrow' soliloquy and the fall require knowing the full arc of Macbeth's crimes and guilt", "mid-excerpt"),
    ("Night, Sections 6", "The death march and liberation section requires knowing Eliezer's family, Auschwitz arrival, and loss of faith", "mid-excerpt"),
    ("In Cold Blood, Parts 1", "Capote's dual narrative needs framing — the Clutter family and the killers' converging stories", "mid-excerpt"),
    ("Great Gatsby, Chapters 5", "Gatsby and Daisy's reunion requires knowing Nick, Gatsby's fabricated persona, Daisy's marriage to Tom, and the green light", "mid-excerpt"),
    ("Great Gatsby, Chapters 8", "Gatsby's death and aftermath require the full arc of his dream and its unraveling", "mid-excerpt"),
    ("Death of a Salesman, Act II", "Willy Loman's confrontation and Requiem require Act I's setup of his delusions, his sons, and the American Dream", "mid-excerpt"),
    ("Their Eyes Were Watching God, Chapters 18", "The hurricane climax is the emotional peak; students need Janie's journey through three marriages and Tea Cake's world", "mid-excerpt"),
    ("Beloved, Part II", "Sethe recognizing Beloved and the 'Sixty Million' section require Part I's ghost-story setup and the revelation of Sethe's past", "mid-excerpt"),
    ("Beloved, Part One, Chapters 1", "While this is Part I, the novel opens in medias res with '124 was spiteful' — students need Reconstruction-era context and the legacy of slavery", "mid-excerpt"),
    ("Invisible Man, Prologue", "The Battle Royal scene is violent and disorienting by design — Jim Crow-era racial dynamics and the narrator's situation must be established", "mid-excerpt"),
    ("As I Lay Dying", "Faulkner's 15 rotating narrators and stream of consciousness are impenetrable without framing the Bundren family and their journey", "mid-excerpt"),
    ("Hamlet, Act II", "Each subsequent act requires previous context; Act II-III drops into feigned madness and 'To be or not to be' mid-arc", "mid-excerpt"),
    ("Hamlet, Act IV", "Ophelia's death and the graveyard scene require knowing the full conspiracy, Polonius's murder, and Hamlet's exile", "mid-excerpt"),
    ("Crucible, Acts III", "Proctor's choice to die for his name requires knowing the Salem accusations, Abigail's manipulation, and Proctor's affair", "mid-excerpt"),
    ("Romeo and Juliet, Act III", "Mercutio and Tybalt's deaths and banishment require knowing the Montague-Capulet feud and the secret marriage", "mid-excerpt"),
    ("Romeo and Juliet, Acts IV", "The resolution and tragedy require the full arc of Acts I-III", "mid-excerpt"),
    ("Things Fall Apart, Part 2", "Okonkwo's exile and the colonial encounter require Part 1's establishment of Igbo culture and Okonkwo's character", "mid-excerpt"),
    ("Slaughterhouse-Five", "Vonnegut's non-linear anti-war narrative blends WWII memoir with science fiction; students need the Dresden bombing context", "mid-excerpt"),
    ("Jane Eyre, Chapters 26", "The interrupted wedding and Jane's declaration require knowing Rochester, Thornfield, and the mystery of Bertha", "mid-excerpt"),
    ("Wuthering Heights, Chapters 9", "Catherine's choice and the death scene require understanding the Heathcliff-Catherine-Edgar triangle and the Yorkshire moors setting", "mid-excerpt"),
    ("Educated, Chapters 10", "University and transformation sections require knowing Tara's survivalist family, lack of formal schooling, and her father's ideology", "mid-excerpt"),
    ("Educated, Chapters 28", "Confrontation and resolution require the full arc of Tara's education and family conflict", "mid-excerpt"),
    ("One Hundred Years of Solitude", "Magical realism; students need to understand the Latin American literary tradition and that the blending of myth and reality is intentional", "mid-excerpt"),
    ("Kite Runner, Chapters 1", "While this opens the novel, students need the pre-Soviet Afghanistan context and Pashtun-Hazara ethnic dynamics for comprehension", "mid-excerpt"),
    ("God of Small Things", "India's caste system and the 'Love Laws' are central; students need the Kerala setting and caste context", "mid-excerpt"),
    ("In the Time of the Butterflies, Part 3", "The reckoning requires knowing the Mirabal sisters, Trujillo's dictatorship, and the resistance movement", "mid-excerpt"),
    ("Warmth of Other Suns, Part 3", "Robert Pershing Foster's story requires knowing the Great Migration's causes and the three migrant narratives Wilkerson traces", "mid-excerpt"),
    ("Color Purple, Letters 1", "Celie's early letters require understanding the rural Jim Crow South and the epistolary novel form", "mid-excerpt"),
    ("Native Son, Book One", "Bigger Thomas's fear requires understanding 1930s Chicago's South Side, the Great Migration, and systemic racial oppression", "mid-excerpt"),
    ("Grapes of Wrath, Chapters 1", "While Chapter 1 opens the novel, the Dust Bowl context and 1930s dispossession of Plains farmers is essential background", "mid-excerpt"),
    ("Handmaid's Tale, Chapters 1", "Gilead's theocratic patriarchy; students need to understand this is a dystopian future America and the Republic's gender-based oppression", "mid-excerpt"),
    ("Brave New World, Chapters 1", "Huxley's pleasure-dystopia with biotechnology and conditioning needs framing — contrast with Orwell's fear-based control", "mid-excerpt"),
    ("Children of Men, Chapter 1", "Near-future infertility premise; students need to know the speculative setup that no children have been born for 25 years", "mid-excerpt"),

    # --- Historical primary sources needing political context ---
    ("Letter from Birmingham Jail", "Students need to know about the Birmingham campaign, why King was jailed, and who the 'white moderate' clergymen were who prompted this letter", "primary-source"),
    ("On Civil Disobedience", "The Mexican-American War and slavery context that provoked Thoreau's essay and his night in jail", "primary-source"),
    ("Common Sense", "1776 colonial context; why monarchy was being questioned and the radical nature of Paine's argument for independence", "primary-source"),
    ("Sinners in the Hands", "Great Awakening context; this is a sermon delivered during a religious revival, not a personal threat — the rhetorical purpose matters", "primary-source"),
    ("Speech After Being Convicted", "Susan B. Anthony's 1872 arrest and trial for the crime of voting while female", "primary-source"),
    ("Federalist No. 10", "The Constitutional Convention debate; what 'factions' meant in 1787 and why Madison feared direct democracy", "primary-source"),
    ("Federalist No. 51", "Separation of powers argument; requires understanding the Articles of Confederation's failure and the ratification debate", "primary-source"),
    ("Ballot or the Bullet", "1964 civil rights context; how Malcolm X's rhetoric and philosophy differed from King's nonviolent approach", "primary-source"),
    ("Checkers Speech", "Nixon's 1952 TV address; students won't understand why this was revolutionary in the history of media-age politics", "primary-source"),
    ("Of Plymouth Plantation", "17th-century Puritan worldview and colonial conditions; Bradford's prose style and religious framework need context", "primary-source"),
    ("Fourth of July", "The devastating irony requires knowing Douglass's audience, the Fugitive Slave Act, and the date (July 5, 1852)", "primary-source"),
    ("Souls of Black Folk", "'Double consciousness' is a precise theoretical concept from 1903, not casual metaphor — Du Bois's framework needs setup", "primary-source"),
    ("Letter from Delano", "Cesar Chavez's farmworker movement and the deliberate rhetorical parallel to King's Birmingham letter", "primary-source"),
    ("Chicago Defender", "Students need to understand the Black press's role in sparking the Great Migration and why these 1916-1920 articles were revolutionary", "primary-source"),
    ("Declaration of Sentiments", "Modeled on the Declaration of Independence; 1848 Seneca Falls Convention context and the women's suffrage movement", "primary-source"),
    ("Narrative of the Life of Frederick Douglass", "The institution of slavery, Douglass's status as an escaped slave writing for an abolitionist audience, and the autobiographical genre's political purpose", "primary-source"),
    ("Incidents in the Life of a Slave Girl", "Harriet Jacobs's gendered experience of slavery; the pseudonym 'Linda Brent' and the political context of publishing a female slave narrative", "primary-source"),
    ("On the Abolition of the English Department", "Ngũgĩ wa Thiong'o's postcolonial argument about language and cultural imperialism needs the Kenyan independence context", "primary-source"),
    ("Nobel Lecture", "Wole Soyinka's significance as the first African Nobel laureate in Literature (1986) and Nigerian political context", "primary-source"),
    ("Autobiography of Malcolm X", "Malcolm X's transformation from street criminal to Nation of Islam leader to global human rights activist — the conversion narrative context", "primary-source"),
]

CATEGORY_LABELS = {
    "satire": "Satire/Irony — Will Be Misread",
    "ancient": "Ancient/Archaic — Major Cultural Distance",
    "mid-excerpt": "Mid-Work Excerpt — Missing Prior Context",
    "primary-source": "Historical Primary Source — Political Context Needed",
    "philosophical": "Philosophical/Experimental — Conceptual Framework Needed",
}


def analyze_contextualization(grade, data):
    """
    Analyze each assessment in a grade for contextual introductions.
    Returns a list of dicts with title, has_context, context_type, reason, first_text.
    """
    results = []
    assessments = get_all_assessments(data)

    for unit_title, assessment in assessments:
        title = get_assessment_title(assessment)

        # Get the first section's stimulus content
        first_text = ""
        tps = assessment.get("test_parts", [])
        if tps:
            secs = tps[0].get("sections", [])
            if secs and secs[0].get("items"):
                stim = secs[0]["items"][0].get("stimulus", {})
                if isinstance(stim, dict):
                    html = stim.get("content_html", "")
                else:
                    html = str(stim)
                first_text = re.sub(r"<[^>]+>", " ", html).strip()
                first_text = re.sub(r"\s+", " ", first_text)

        # Remove the title from the beginning to get actual content start
        text_after_title = first_text
        title_lower = title.lower()
        if title_lower in first_text[:300].lower():
            idx = first_text.lower().find(title_lower)
            text_after_title = first_text[idx + len(title):].strip()

        # Check if this text has a contextual introduction
        check_region = text_after_title[:500].lower()
        intro_phrases = [
            "in this excerpt", "this passage", "this story takes place",
            "background:", "context:", "about the author", "about this",
            "introduction:", "before you read", "before reading",
            "in this selection", "historical context", "cultural context",
            "you are about to read", "as you read, consider",
            "the following excerpt is from", "the following passage is from",
        ]
        has_context = False
        context_note = ""
        for phrase in intro_phrases:
            if phrase in check_region:
                has_context = True
                context_note = f"Contains '{phrase}' framing"
                break

        # Check if text matches any entry in TEXTS_NEEDING_CONTEXT
        needs_context = False
        reason = ""
        category = ""
        for frag, r, cat in TEXTS_NEEDING_CONTEXT:
            if frag.lower() in title.lower():
                needs_context = True
                reason = r
                category = cat
                break

        if needs_context:
            results.append({
                "title": title,
                "has_context": has_context,
                "context_note": context_note,
                "needs_context": True,
                "reason": reason,
                "category": category,
                "first_200": text_after_title[:200],
            })

    return results


# ============================================================================
# HTML REPORT GENERATION
# ============================================================================

def pct_bar(pct, bg_class, label=""):
    """Generate a Tailwind-styled bar for a percentage."""
    return (
        f'<div class="flex items-center my-1">'
        f'<div class="w-[60px] text-right pr-2 text-[13px] text-gray-500">{label}</div>'
        f'<div class="flex-1 bg-gray-200 rounded h-[22px] relative">'
        f'<div class="h-full rounded min-w-[2px] transition-all duration-300 {bg_class}" style="width:{min(pct, 100):.1f}%;"></div>'
        f'</div>'
        f'<div class="w-[55px] text-right pl-2 text-[13px] font-semibold">{pct:.1f}%</div>'
        f'</div>'
    )


def generate_html(all_grades, context_data=None):
    """Generate the complete HTML report."""
    # Tailwind class constants for tables
    TH = 'class="bg-gray-50 px-3 py-2.5 text-left font-semibold text-indigo-900 border-b-2 border-gray-300"'
    TD = 'class="px-3 py-2 border-b border-gray-200"'
    TABLE = 'class="w-full border-collapse mt-3 text-sm"'
    SUMMARY_CLS = 'class="cursor-pointer text-xs text-blue-600 font-medium hover:underline"'

    grade_tabs = ""
    grade_panels = ""

    for grade in sorted(all_grades.keys()):
        m = all_grades[grade]
        active_class = " active" if grade == 3 else ""
        active_display = "block" if grade == 3 else "none"

        tab_active_classes = " text-blue-600 border-blue-600 font-semibold" if grade == 3 else ""
        grade_tabs += f'<button class="tab-btn px-4 py-3 border-b-[3px] border-transparent text-sm font-medium text-gray-500 hover:text-blue-600 hover:bg-blue-50 cursor-pointer whitespace-nowrap transition-all{tab_active_classes}" onclick="showGrade({grade}, this)">Grade {grade}</button>\n'

        # --- Build grade panel ---
        panel = f'<div id="grade-{grade}" class="grade-panel" style="display:{active_display};">\n'

        # Viewer link
        if grade <= 8:
            viewer_url = f"https://alexwrega.github.io/article-viewer/#grade={grade}"
        else:
            viewer_url = f"https://alexwrega.github.io/article-viewer-9-12/#grade={grade}"

        # Summary
        panel += f'<div class="bg-white rounded-lg p-6 mb-5 shadow-sm">\n'
        panel += f'<h2 class="text-lg font-semibold text-indigo-900 mb-4 pb-2 border-b-2 border-gray-200">Grade {grade} — Overview <a href="{viewer_url}" target="_blank" class="ml-2 text-sm font-normal text-blue-500 hover:text-blue-700 hover:underline">View articles &rarr;</a></h2>\n'
        panel += f'<div class="grid grid-cols-[repeat(auto-fit,minmax(140px,1fr))] gap-4">\n'
        panel += f'<div class="text-center p-4 bg-gray-50 rounded-lg"><div class="text-3xl font-bold text-blue-600">{m["total_articles"]}</div><div class="text-xs text-gray-500 mt-1">Articles</div></div>\n'
        panel += f'<div class="text-center p-4 bg-gray-50 rounded-lg"><div class="text-3xl font-bold text-blue-600">{m["total_questions"]}</div><div class="text-xs text-gray-500 mt-1">Questions</div></div>\n'
        panel += f'<div class="text-center p-4 bg-gray-50 rounded-lg"><div class="text-3xl font-bold text-blue-600">{m["num_units"]}</div><div class="text-xs text-gray-500 mt-1">Units</div></div>\n'
        if m["lexile_values"]:
            avg_lex = sum(m["lexile_values"]) / len(m["lexile_values"])
            min_lex = min(m["lexile_values"])
            max_lex = max(m["lexile_values"])
            panel += f'<div class="text-center p-4 bg-gray-50 rounded-lg"><div class="text-3xl font-bold text-blue-600">{avg_lex:.0f}</div><div class="text-xs text-gray-500 mt-1">Avg Lexile ({min_lex}–{max_lex})</div></div>\n'
        else:
            pl_info = PLANNED_LEXILE.get(grade)
            if pl_info:
                panel += f'<div class="text-center p-4 bg-gray-50 rounded-lg"><div class="text-3xl font-bold text-pink-600">{pl_info["midpoint"]}</div><div class="text-xs text-gray-500 mt-1">Planned Lexile ({pl_info["range"]})</div></div>\n'
            else:
                panel += f'<div class="text-center p-4 bg-gray-50 rounded-lg"><div class="text-3xl font-bold text-pink-600">N/A</div><div class="text-xs text-gray-500 mt-1">Lexile (no data)</div></div>\n'
        panel += f'</div></div>\n'

        # Curriculum Plan vs Reality (grades 9-12 only)
        if m.get("curriculum_plan") and m.get("plan_crossref"):
            cp = m["curriculum_plan"]
            xr = m["plan_crossref"]
            pl = PLANNED_LEXILE.get(grade, {})
            pc = PLANNED_COUNTS.get(grade, {})

            panel += f'<div class="bg-white rounded-lg p-6 mb-5 shadow-sm">\n'
            panel += f'<h2 class="text-lg font-semibold text-indigo-900 mb-4 pb-2 border-b-2 border-gray-200">Curriculum Plan vs. Current Content</h2>\n'
            panel += f'<p class="text-xs text-gray-500 mb-3 italic">Cross-referencing the high school curriculum plan (XLSX) against '
            panel += f'actual QTI assessment data. The plan specifies {cp["total"]} texts for grade {grade} '
            panel += f'({cp["custom_count"]} are custom texts that need to be commissioned).</p>\n'

            # Plan overview stats
            panel += f'<div class="grid grid-cols-[repeat(auto-fit,minmax(140px,1fr))] gap-4">\n'
            panel += f'<div class="text-center p-4 bg-gray-50 rounded-lg"><div class="text-3xl font-bold text-blue-600">{cp["total"]}</div><div class="text-xs text-gray-500 mt-1">Planned Texts</div></div>\n'
            panel += f'<div class="text-center p-4 bg-gray-50 rounded-lg"><div class="text-3xl font-bold text-blue-600">{m["total_articles"]}</div><div class="text-xs text-gray-500 mt-1">Current QTI Articles</div></div>\n'
            panel += f'<div class="text-center p-4 bg-gray-50 rounded-lg"><div class="text-3xl font-bold text-blue-600">{len(xr["found_in_qti"])}</div><div class="text-xs text-gray-500 mt-1">Matched to Plan</div></div>\n'
            panel += f'<div class="text-center p-4 bg-gray-50 rounded-lg"><div class="text-3xl font-bold text-blue-600">{len(xr["extra_in_qti"])}</div><div class="text-xs text-gray-500 mt-1">Not in Plan</div></div>\n'
            panel += f'</div>\n'

            # Planned Lexile info
            if pl:
                panel += f'<div class="my-4 bg-blue-50 border-l-4 border-blue-500 p-3 rounded-md">\n'
                panel += f'<strong>Planned Lexile Range:</strong> {pl["range"]} (midpoint: {pl["midpoint"]}L)<br>\n'
                panel += f'<strong>Scaffolded:</strong> Early {pl["early"]} → Mid {pl["mid"]} → Late {pl["late"]}<br>\n'
                panel += f'<strong>Planned Word Count:</strong> {PLANNED_WORD_COUNTS.get(grade, "N/A")} per article\n'
                panel += f'</div>\n'

            # Plan vs reality comparison table
            panel += f'<table class="w-full border-collapse mt-3 text-sm"><thead><tr><th class="bg-gray-50 px-3 py-2.5 text-left font-semibold text-indigo-900 border-b-2 border-gray-300">Metric</th><th class="bg-gray-50 px-3 py-2.5 text-left font-semibold text-indigo-900 border-b-2 border-gray-300">Plan</th><th class="bg-gray-50 px-3 py-2.5 text-left font-semibold text-indigo-900 border-b-2 border-gray-300">Current QTI</th><th class="bg-gray-50 px-3 py-2.5 text-left font-semibold text-indigo-900 border-b-2 border-gray-300">Status</th></tr></thead><tbody>\n'
            panel += f'<tr><td class="px-3 py-2 border-b border-gray-200">Total Articles</td><td class="px-3 py-2 border-b border-gray-200">{pc.get("articles", "?")}</td><td class="px-3 py-2 border-b border-gray-200">{m["total_articles"]}</td>'
            diff = m["total_articles"] - pc.get("articles", 0)
            if abs(diff) < 5:
                panel += f'<td class="px-3 py-2 border-b border-gray-200 text-green-600">✓ Close match</td></tr>\n'
            else:
                panel += f'<td class="px-3 py-2 border-b border-gray-200 text-red-600">⚠ {diff:+d} difference</td></tr>\n'

            panel += f'<tr><td class="px-3 py-2 border-b border-gray-200">Units</td><td class="px-3 py-2 border-b border-gray-200">{pc.get("units", "?")}</td><td class="px-3 py-2 border-b border-gray-200">{m["num_units"]}</td>'
            if m["num_units"] < 3:
                panel += f'<td class="px-3 py-2 border-b border-gray-200 text-red-600">⚠ No thematic units in QTI</td></tr>\n'
            else:
                panel += f'<td class="px-3 py-2 border-b border-gray-200">—</td></tr>\n'

            panel += f'<tr><td class="px-3 py-2 border-b border-gray-200">Literary Texts</td><td class="px-3 py-2 border-b border-gray-200">{pc.get("literary", "?")}</td><td class="px-3 py-2 border-b border-gray-200">{m["literary_count"]}</td>'
            panel += f'<td class="px-3 py-2 border-b border-gray-200">—</td></tr>\n'
            panel += f'<tr><td class="px-3 py-2 border-b border-gray-200">Informational Texts</td><td class="px-3 py-2 border-b border-gray-200">{pc.get("informational", "?")}</td><td class="px-3 py-2 border-b border-gray-200">{m["informational_count"]}</td>'
            panel += f'<td class="px-3 py-2 border-b border-gray-200">—</td></tr>\n'

            panel += f'<tr><td class="px-3 py-2 border-b border-gray-200">Lexile Data</td><td class="px-3 py-2 border-b border-gray-200">{pl.get("range", "?")}</td>'
            if m["lexile_values"]:
                panel += f'<td class="px-3 py-2 border-b border-gray-200">{min(m["lexile_values"])}L–{max(m["lexile_values"])}L</td><td class="px-3 py-2 border-b border-gray-200">—</td></tr>\n'
            else:
                panel += f'<td class="px-3 py-2 border-b border-gray-200 text-red-600">No data (all 0)</td><td class="px-3 py-2 border-b border-gray-200 text-red-600">⚠ Cannot verify complexity</td></tr>\n'

            panel += f'<tr><td class="px-3 py-2 border-b border-gray-200">Custom Texts Needed</td><td colspan="2" class="px-3 py-2 border-b border-gray-200">{cp["custom_count"]} texts need to be written</td>'
            panel += f'<td class="px-3 py-2 border-b border-gray-200 text-pink-500">⚠ Not yet created</td></tr>\n'
            panel += f'</tbody></table>\n'

            # Planned unit structure
            if cp["units"]:
                panel += f'<h3 class="mt-4 text-indigo-900 font-semibold text-base">Planned Unit Structure</h3>\n'
                panel += f'<ol class="pl-6">\n'
                for u in cp["units"]:
                    panel += f'<li class="my-1 text-sm">{escape(u)}</li>\n'
                panel += f'</ol>\n'
                if m["num_units"] <= 1:
                    panel += f'<div class="p-3 rounded-md mb-3 text-sm font-medium bg-red-50 text-red-800 border-l-4 border-red-500">The QTI data has no thematic unit structure — all {m["total_articles"]} articles are in a single "Reading Exercises" unit, unlike the {len(cp["units"])} well-organized units specified in the curriculum plan.</div>\n'

            # Planned genre distribution
            if cp.get("genres"):
                panel += f'<h3 class="mt-4 text-indigo-900 font-semibold text-base">Planned Genre Distribution</h3>\n'
                panel += f'<table class="w-full border-collapse mt-3 text-sm"><thead><tr><th class="bg-gray-50 px-3 py-2.5 text-left font-semibold text-indigo-900 border-b-2 border-gray-300">Genre</th><th class="bg-gray-50 px-3 py-2.5 text-left font-semibold text-indigo-900 border-b-2 border-gray-300">Count</th><th class="bg-gray-50 px-3 py-2.5 text-left font-semibold text-indigo-900 border-b-2 border-gray-300">% of Plan</th></tr></thead><tbody>\n'
                for genre, count in sorted(cp["genres"].items(), key=lambda x: -x[1]):
                    gpct = count / cp["total"] * 100 if cp["total"] else 0
                    panel += f'<tr class="hover:bg-gray-50"><td class="px-3 py-2 border-b border-gray-200">{escape(genre)}</td><td class="px-3 py-2 border-b border-gray-200">{count}</td><td class="px-3 py-2 border-b border-gray-200">{gpct:.1f}%</td></tr>\n'
                panel += f'</tbody></table>\n'

            # Missing texts from plan
            if xr["missing_from_qti"]:
                panel += f'<h3 class="mt-4 text-indigo-900 font-semibold text-base">Planned Texts Missing from QTI ({len(xr["missing_from_qti"])})</h3>\n'
                panel += f'<p class="text-xs text-gray-500 mb-3 italic">Published texts specified in the curriculum plan that are not found in the current QTI data.</p>\n'
                panel += f'<details open><summary class="cursor-pointer text-xs text-blue-600 font-medium hover:underline">Missing published texts ({len(xr["missing_from_qti"])})</summary>\n'
                panel += f'<table class="w-full border-collapse mt-3 text-sm"><thead><tr><th class="bg-gray-50 px-3 py-2.5 text-left font-semibold text-indigo-900 border-b-2 border-gray-300">Title</th><th class="bg-gray-50 px-3 py-2.5 text-left font-semibold text-indigo-900 border-b-2 border-gray-300">Author</th><th class="bg-gray-50 px-3 py-2.5 text-left font-semibold text-indigo-900 border-b-2 border-gray-300">Type</th><th class="bg-gray-50 px-3 py-2.5 text-left font-semibold text-indigo-900 border-b-2 border-gray-300">Unit</th></tr></thead><tbody>\n'
                for t in xr["missing_from_qti"][:50]:
                    panel += f'<tr class="hover:bg-gray-50"><td class="px-3 py-2 border-b border-gray-200">{escape(t["title"][:80])}</td><td class="px-3 py-2 border-b border-gray-200">{escape(t["author"][:40])}</td>'
                    panel += f'<td class="px-3 py-2 border-b border-gray-200">{escape(t["text_type"][:40])}</td><td class="px-3 py-2 border-b border-gray-200 text-xs">{escape(t["unit"][:50])}</td></tr>\n'
                if len(xr["missing_from_qti"]) > 50:
                    panel += f'<tr><td colspan="4" class="px-3 py-2 border-b border-gray-200 text-gray-500 italic">... and {len(xr["missing_from_qti"]) - 50} more</td></tr>\n'
                panel += f'</tbody></table></details>\n'

            # Custom texts not yet written
            if xr["custom_not_yet_written"]:
                panel += f'<h3 class="mt-4 text-indigo-900 font-semibold text-base">Custom Texts Not Yet Written ({len(xr["custom_not_yet_written"])})</h3>\n'
                panel += f'<p class="text-xs text-gray-500 mb-3 italic">The plan calls for custom informational texts to be commissioned. These are not yet in the QTI data.</p>\n'
                panel += f'<details><summary class="cursor-pointer text-xs text-blue-600 font-medium hover:underline">Custom texts needed ({len(xr["custom_not_yet_written"])})</summary>\n'
                panel += f'<table class="w-full border-collapse mt-3 text-sm"><thead><tr><th class="bg-gray-50 px-3 py-2.5 text-left font-semibold text-indigo-900 border-b-2 border-gray-300">Title</th><th class="bg-gray-50 px-3 py-2.5 text-left font-semibold text-indigo-900 border-b-2 border-gray-300">Unit</th><th class="bg-gray-50 px-3 py-2.5 text-left font-semibold text-indigo-900 border-b-2 border-gray-300">Notes</th></tr></thead><tbody>\n'
                for t in xr["custom_not_yet_written"][:40]:
                    panel += f'<tr class="hover:bg-gray-50"><td class="px-3 py-2 border-b border-gray-200">{escape(t["title"][:80])}</td>'
                    panel += f'<td class="px-3 py-2 border-b border-gray-200 text-xs">{escape(t["unit"][:50])}</td>'
                    panel += f'<td class="px-3 py-2 border-b border-gray-200 text-xs">{escape(t["notes"][:150])}</td></tr>\n'
                panel += f'</tbody></table></details>\n'

            # Extra QTI texts not in plan
            if xr["extra_in_qti"]:
                panel += f'<h3 class="mt-4 text-indigo-900 font-semibold text-base">QTI Texts Not in Plan ({len(xr["extra_in_qti"])})</h3>\n'
                panel += f'<p class="text-xs text-gray-500 mb-3 italic">Articles in current QTI data that are NOT in the curriculum plan.</p>\n'
                panel += f'<details><summary class="cursor-pointer text-xs text-blue-600 font-medium hover:underline">Extra QTI texts ({len(xr["extra_in_qti"])})</summary><ul class="pl-6 max-h-[300px] overflow-y-auto">\n'
                for t in sorted(xr["extra_in_qti"])[:80]:
                    panel += f'<li class="text-[13px] my-0.5 text-gray-500">{escape(t)}</li>\n'
                if len(xr["extra_in_qti"]) > 80:
                    panel += f'<li class="text-gray-500 italic">... and {len(xr["extra_in_qti"]) - 80} more</li>\n'
                panel += f'</ul></details>\n'

            # Match summary
            panel += f'<div class="mt-4 bg-sky-50 border-l-4 border-sky-600 p-3 rounded-md">\n'
            panel += f'<strong>Plan Match Rate:</strong> {xr["match_rate"]:.0f}% of planned texts found in QTI data '
            panel += f'({len(xr["found_in_qti"])} of {cp["total"]}). '
            panel += f'{len(xr["missing_from_qti"])} published texts are missing, '
            panel += f'{cp["custom_count"]} custom texts need to be written, '
            panel += f'and {len(xr["extra_in_qti"])} current QTI articles are not in the plan.\n'
            panel += f'</div>\n'

            panel += f'</div>\n'

        # Literary vs Informational
        panel += f'<div class="bg-white rounded-lg p-6 mb-5 shadow-sm">\n'
        panel += f'<h2 class="text-lg font-semibold text-indigo-900 mb-4 pb-2 border-b-2 border-gray-200">Literary vs. Informational Texts</h2>\n'
        panel += f'<p class="text-xs text-gray-500 mb-3 italic">Literary = the text IS a literary work (poem, story, play, myth). '
        panel += f'Informational = everything else, including texts <em>about</em> literary works.</p>\n'
        panel += pct_bar(m["literary_pct"], "bg-blue-400", "Literary")
        panel += pct_bar(m["informational_pct"], "bg-red-400", "Info")
        panel += f'<table {TABLE}><thead><tr><th {TH}>Type</th><th {TH}>Count</th><th {TH}>Percentage</th></tr></thead><tbody>\n'
        panel += f'<tr class="hover:bg-gray-50"><td {TD}>Literary</td><td {TD}>{m["literary_count"]}</td><td {TD}>{m["literary_pct"]:.1f}%</td></tr>\n'
        panel += f'<tr class="hover:bg-gray-50"><td {TD}>Informational</td><td {TD}>{m["informational_count"]}</td><td {TD}>{m["informational_pct"]:.1f}%</td></tr>\n'
        panel += f'</tbody></table>\n'

        # Collapsible literary titles
        if m["literary_titles"]:
            panel += f'<details><summary {SUMMARY_CLS}>Literary texts ({len(m["literary_titles"])})</summary><ul class="pl-6 max-h-[300px] overflow-y-auto">\n'
            for t in sorted(m["literary_titles"]):
                panel += f'<li class="text-[13px] my-0.5 text-gray-500">{escape(t)}</li>\n'
            panel += f'</ul></details>\n'

        panel += f'</div>\n'

        # Original Text vs Synopsis (grades 3-8 only)
        if grade <= 8 and m["total_literary_unit_assessments"] > 0:
            total_lit = m["total_literary_unit_assessments"]
            orig_n = m["originality_counts"].get("original", 0)
            hybrid_n = m["originality_counts"].get("synopsis_with_quotes", 0)
            syn_n = m["originality_counts"].get("synopsis", 0)
            orig_pct = orig_n / total_lit * 100 if total_lit else 0
            hybrid_pct = hybrid_n / total_lit * 100 if total_lit else 0
            syn_pct = syn_n / total_lit * 100 if total_lit else 0

            panel += f'<div class="bg-white rounded-lg p-6 mb-5 shadow-sm">\n'
            panel += f'<h2 class="text-lg font-semibold text-indigo-900 mb-4 pb-2 border-b-2 border-gray-200">Original Text vs. Synopsis (Literary Units)</h2>\n'
            panel += f'<p class="text-xs text-gray-500 mb-3 italic">Checks whether assessments in literary units present the '
            panel += f'actual original literary work or a curriculum-written synopsis/retelling/educational article about it.</p>\n'

            panel += pct_bar(orig_pct, "bg-emerald-400", "Original")
            panel += pct_bar(hybrid_pct, "bg-cyan-400", "Hybrid")
            panel += pct_bar(syn_pct, "bg-red-400", "Synopsis")

            panel += f'<table {TABLE}><thead><tr><th {TH}>Classification</th><th {TH}>Count</th><th {TH}>Percentage</th><th {TH}>Description</th></tr></thead><tbody>\n'
            panel += f'<tr class="hover:bg-gray-50"><td {TD}><strong>Original Text</strong></td><td {TD}>{orig_n}</td><td {TD}>{orig_pct:.1f}%</td>'
            panel += f'<td {TD}>Presents actual original literary text as the primary reading</td></tr>\n'
            panel += f'<tr class="hover:bg-gray-50"><td {TD}><strong>Synopsis + Quotes</strong></td><td {TD}>{hybrid_n}</td><td {TD}>{hybrid_pct:.1f}%</td>'
            panel += f'<td {TD}>Curriculum-written article with embedded original excerpts</td></tr>\n'
            panel += f'<tr class="hover:bg-gray-50"><td {TD}><strong>Pure Synopsis</strong></td><td {TD}>{syn_n}</td><td {TD}>{syn_pct:.1f}%</td>'
            panel += f'<td {TD}>Entirely curriculum-written synopsis, retelling, or educational article</td></tr>\n'
            panel += f'</tbody></table>\n'

            # Collapsible details
            for cat, label, details_list in [
                ("original", "Original text assessments", m["originality_details"]["original"]),
                ("synopsis_with_quotes", "Synopsis + quotes assessments", m["originality_details"]["synopsis_with_quotes"]),
                ("synopsis", "Pure synopsis assessments", m["originality_details"]["synopsis"]),
            ]:
                if details_list:
                    panel += f'<details><summary {SUMMARY_CLS}>{label} ({len(details_list)})</summary><ul class="pl-6 max-h-[300px] overflow-y-auto">\n'
                    for t in sorted(details_list):
                        panel += f'<li class="text-[13px] my-0.5 text-gray-500">{escape(t)}</li>\n'
                    panel += f'</ul></details>\n'

            panel += f'</div>\n'

        # Excerpt vs Non-excerpt
        panel += f'<div class="bg-white rounded-lg p-6 mb-5 shadow-sm">\n'
        panel += f'<h2 class="text-lg font-semibold text-indigo-900 mb-4 pb-2 border-b-2 border-gray-200">Excerpt vs. Non-Excerpt</h2>\n'
        panel += f'<p class="text-xs text-gray-500 mb-3 italic">Excerpt = a portion of a longer literary work. Non-excerpt = complete or self-contained text.</p>\n'
        panel += pct_bar(m["excerpt_pct"], "bg-red-400", "Excerpt")
        panel += pct_bar(m["non_excerpt_pct"], "bg-blue-400", "Complete")
        panel += f'<table {TABLE}><thead><tr><th {TH}>Type</th><th {TH}>Count</th><th {TH}>Percentage</th></tr></thead><tbody>\n'
        panel += f'<tr class="hover:bg-gray-50"><td {TD}>Excerpt</td><td {TD}>{m["excerpt_count"]}</td><td {TD}>{m["excerpt_pct"]:.1f}%</td></tr>\n'
        panel += f'<tr class="hover:bg-gray-50"><td {TD}>Non-Excerpt</td><td {TD}>{m["non_excerpt_count"]}</td><td {TD}>{m["non_excerpt_pct"]:.1f}%</td></tr>\n'
        panel += f'</tbody></table>\n'

        if m["excerpt_titles"]:
            panel += f'<details><summary {SUMMARY_CLS}>Excerpted texts ({len(m["excerpt_titles"])})</summary><ul class="pl-6 max-h-[300px] overflow-y-auto">\n'
            for t in sorted(m["excerpt_titles"]):
                panel += f'<li class="text-[13px] my-0.5 text-gray-500">{escape(t)}</li>\n'
            panel += f'</ul></details>\n'

        panel += f'</div>\n'

        # Standards Coverage — split into RL and RI
        panel += f'<div class="bg-white rounded-lg p-6 mb-5 shadow-sm">\n'
        panel += f'<h2 class="text-lg font-semibold text-indigo-900 mb-4 pb-2 border-b-2 border-gray-200">Standards Coverage (CCSS ELA Reading)</h2>\n'
        panel += f'<p class="text-xs text-gray-500 mb-3 italic">Coverage estimated by keyword analysis of question prompts and unit titles, '
        panel += f'split by text type. RL = Reading Literature, RI = Reading Informational Text. '
        panel += f'Standard 8 applies to informational text only. '
        panel += f'Standard 10 (Range of Reading) is assessed via Lexile coverage.</p>\n'

        g = grade_band(grade)

        # --- RL Table ---
        panel += f'<h3 class="mt-4 text-indigo-900 font-semibold text-base">RL — Reading Literature</h3>\n'
        panel += f'<table {TABLE}><thead><tr>'
        panel += f'<th {TH}>Standard</th><th {TH}>Description</th><th {TH}>Questions</th><th {TH}>% of Questions</th><th {TH}>Assessments</th></tr></thead><tbody>\n'

        std_row_map = {"std-none": "bg-red-200 text-red-900", "std-weak": "bg-sky-200 text-sky-800", "std-moderate": "", "std-strong": "bg-green-200 text-green-900"}

        for std in ["R1", "R2", "R3", "R4", "R5", "R6", "R7", "R9", "R10"]:
            s = m["standards"][std]
            rl_q = s["rl_questions"]
            rl_pct = (rl_q / m["total_questions"] * 100) if m["total_questions"] else 0
            rl_a = s["rl_assessments"]
            label = ccss_label(grade, std, "RL")

            if std == "R10":
                row_class = std_row_map["std-moderate"] if m["lexile_values"] else std_row_map["std-none"]
                if m["lexile_values"]:
                    avg_lex = sum(m["lexile_values"]) / len(m["lexile_values"])
                    r10_note = f'Lexile range {min(m["lexile_values"])}–{max(m["lexile_values"])} (avg {avg_lex:.0f})'
                else:
                    r10_note = "No Lexile data"
                panel += f'<tr class="{row_class}"><td {TD}><strong>{escape(label)}</strong></td>'
                panel += f'<td {TD}>{escape(STANDARD_DESCRIPTIONS[std])}</td>'
                panel += f'<td colspan="3" {TD}><span class="italic">{escape(r10_note)}</span></td></tr>\n'
            else:
                if rl_q == 0:
                    row_class = std_row_map["std-none"]
                elif rl_pct < 3:
                    row_class = std_row_map["std-weak"]
                elif rl_pct < 10:
                    row_class = std_row_map["std-moderate"]
                else:
                    row_class = std_row_map["std-strong"]

                # RL-specific descriptions for standards that differ
                rl_desc = STANDARD_DESCRIPTIONS[std]
                if std == "R3":
                    rl_desc = "Analyze how characters develop, interact, and advance plot"
                elif std == "R9":
                    rl_desc = "Compare treatment of themes/topics across genres and periods"

                panel += f'<tr class="{row_class}"><td {TD}><strong>{escape(label)}</strong></td><td {TD}>{escape(rl_desc)}</td>'
                panel += f'<td {TD}>{rl_q}</td><td {TD}>{rl_pct:.1f}%</td><td {TD}>{rl_a}</td></tr>\n'

        panel += f'</tbody></table>\n'

        # --- RI Table ---
        panel += f'<h3 class="mt-5 text-indigo-900 font-semibold text-base">RI — Reading Informational Text</h3>\n'
        panel += f'<table {TABLE}><thead><tr>'
        panel += f'<th {TH}>Standard</th><th {TH}>Description</th><th {TH}>Questions</th><th {TH}>% of Questions</th><th {TH}>Assessments</th></tr></thead><tbody>\n'

        for std in ["R1", "R2", "R3", "R4", "R5", "R6", "R7", "R8", "R9", "R10"]:
            s = m["standards"][std]
            ri_q = s["ri_questions"]
            ri_pct = (ri_q / m["total_questions"] * 100) if m["total_questions"] else 0
            ri_a = s["ri_assessments"]
            label = ccss_label(grade, std, "RI")

            if std == "R10":
                row_class = std_row_map["std-moderate"] if m["lexile_values"] else std_row_map["std-none"]
                if m["lexile_values"]:
                    avg_lex = sum(m["lexile_values"]) / len(m["lexile_values"])
                    r10_note = f'Lexile range {min(m["lexile_values"])}–{max(m["lexile_values"])} (avg {avg_lex:.0f})'
                else:
                    r10_note = "No Lexile data"
                panel += f'<tr class="{row_class}"><td {TD}><strong>{escape(label)}</strong></td>'
                panel += f'<td {TD}>{escape(STANDARD_DESCRIPTIONS[std])}</td>'
                panel += f'<td colspan="3" {TD}><span class="italic">{escape(r10_note)}</span></td></tr>\n'
            else:
                if ri_q == 0:
                    row_class = std_row_map["std-none"]
                elif ri_pct < 3:
                    row_class = std_row_map["std-weak"]
                elif ri_pct < 10:
                    row_class = std_row_map["std-moderate"]
                else:
                    row_class = std_row_map["std-strong"]

                # RI-specific descriptions
                ri_desc = STANDARD_DESCRIPTIONS[std]
                if std == "R3":
                    ri_desc = "Analyze how individuals, events, and ideas develop and interact"
                elif std == "R9":
                    ri_desc = "Compare/contrast how authors present information on same topic"

                panel += f'<tr class="{row_class}"><td {TD}><strong>{escape(label)}</strong></td><td {TD}>{escape(ri_desc)}</td>'
                panel += f'<td {TD}>{ri_q}</td><td {TD}>{ri_pct:.1f}%</td><td {TD}>{ri_a}</td></tr>\n'

        panel += f'</tbody></table>\n'
        panel += f'</div>\n'

        # Answer Distribution
        panel += f'<div class="bg-white rounded-lg p-6 mb-5 shadow-sm">\n'
        panel += f'<h2 class="text-lg font-semibold text-indigo-900 mb-4 pb-2 border-b-2 border-gray-200">Correct Answer Distribution</h2>\n'
        panel += f'<p class="text-xs text-gray-500 mb-3 italic">Ideal distribution: ~25% each for A, B, C, D.</p>\n'

        labels = ["A", "B", "C", "D"]
        colors = ["bg-blue-400", "bg-emerald-400", "bg-cyan-400", "bg-red-400"]
        for i in range(4):
            count = m["answer_distribution"].get(i, 0)
            pct = (count / m["answer_total"] * 100) if m["answer_total"] else 0
            panel += pct_bar(pct, colors[i], labels[i])

        panel += f'<table {TABLE}><thead><tr><th {TH}>Answer</th><th {TH}>Count</th><th {TH}>Percentage</th><th {TH}>Deviation from 25%</th></tr></thead><tbody>\n'
        dev_classes = {"dev-bad": "text-red-600 font-semibold", "dev-warn": "text-pink-600 font-medium", "dev-ok": "text-green-700"}
        for i in range(4):
            count = m["answer_distribution"].get(i, 0)
            pct = (count / m["answer_total"] * 100) if m["answer_total"] else 0
            dev = pct - 25
            dev_class = dev_classes["dev-bad"] if abs(dev) > 10 else (dev_classes["dev-warn"] if abs(dev) > 5 else dev_classes["dev-ok"])
            panel += f'<tr class="hover:bg-gray-50"><td {TD}><strong>{labels[i]}</strong></td><td {TD}>{count}</td><td {TD}>{pct:.1f}%</td>'
            panel += f'<td class="px-3 py-2 border-b border-gray-200 {dev_class}">{dev:+.1f}%</td></tr>\n'
        panel += f'</tbody></table>\n'
        panel += f'</div>\n'

        # Flagged Items
        if grade >= 5:
            panel += f'<div class="bg-white rounded-lg p-6 mb-5 shadow-sm">\n'
            panel += f'<h2 class="text-lg font-semibold text-indigo-900 mb-4 pb-2 border-b-2 border-gray-200">Flagged Content (from Appropriateness Audit)</h2>\n'
            panel += f'<p class="text-xs text-gray-500 mb-3 italic">Items previously flagged at '
            panel += f'<a href="https://ilmych.github.io/reading-appropriateness/" target="_blank">'
            panel += f'ilmych.github.io/reading-appropriateness</a>. '
            panel += f'Checked against current grade {grade} content.</p>\n'

            if m["flagged_items"]:
                panel += f'<div class="p-3 rounded-md mb-3 text-sm font-medium bg-red-50 text-red-800 border-l-4 border-red-500">⚠ {len(m["flagged_items"])} flagged item(s) found in current content</div>\n'
                panel += f'<table class="w-full border-collapse mt-3 text-[13px]"><thead><tr>'
                panel += f'<th {TH}>Flagged Title</th><th {TH}>Current Title</th><th {TH}>Prior Status</th><th {TH}>Severity</th><th {TH}>Flags</th><th {TH}>Match</th></tr></thead><tbody>\n'
                for item in m["flagged_items"]:
                    sev_class = "text-red-800 font-bold" if item["severity"] == "HIGH" else "text-pink-600 font-semibold"
                    status_class = "text-gray-500" if item["status"] == "Used" else "text-red-800 font-semibold"
                    panel += f'<tr class="hover:bg-gray-50"><td class="px-3 py-2 border-b border-gray-200 align-top">{escape(item["flagged_title"])}</td>'
                    panel += f'<td class="px-3 py-2 border-b border-gray-200 align-top">{escape(item["current_title"])}</td>'
                    panel += f'<td class="px-3 py-2 border-b border-gray-200 align-top {status_class}">{item["status"]}</td>'
                    panel += f'<td class="px-3 py-2 border-b border-gray-200 align-top {sev_class}">{item["severity"]}</td>'
                    panel += f'<td class="px-3 py-2 border-b border-gray-200 align-top text-xs max-w-[250px]">{escape(item["flags"])}</td>'
                    panel += f'<td class="px-3 py-2 border-b border-gray-200 align-top">{item["match_type"]}</td></tr>\n'
                panel += f'</tbody></table>\n'
            else:
                panel += f'<div class="p-3 rounded-md mb-3 text-sm font-medium bg-green-50 text-green-800 border-l-4 border-green-500">No flagged items found in current content.</div>\n'
            panel += f'</div>\n'

        # Strengths
        panel += f'<div class="bg-white rounded-lg p-6 mb-5 shadow-sm">\n'
        panel += f'<h2 class="text-lg font-semibold text-indigo-900 mb-4 pb-2 border-b-2 border-gray-200">Strengths</h2>\n'
        panel += f'<ul class="list-none">\n'
        for s in m["strengths"]:
            panel += f'<li class="my-2 pl-2 text-green-800 before:content-[\'\\2713_\'] before:font-bold before:text-green-600">{escape(s)}</li>\n'
        panel += f'</ul></div>\n'

        # Weaknesses
        panel += f'<div class="bg-white rounded-lg p-6 mb-5 shadow-sm">\n'
        panel += f'<h2 class="text-lg font-semibold text-indigo-900 mb-4 pb-2 border-b-2 border-gray-200">Weaknesses</h2>\n'
        panel += f'<ul class="list-none">\n'
        for w in m["weaknesses"]:
            panel += f'<li class="my-2.5 pl-2 text-red-900 leading-normal before:content-[\'\\2717_\'] before:font-bold before:text-red-600">{escape(w)}</li>\n'
        panel += f'</ul></div>\n'

        # Unit list
        if m["unit_titles"]:
            panel += f'<div class="bg-white rounded-lg p-6 mb-5 shadow-sm">\n'
            panel += f'<h2 class="text-lg font-semibold text-indigo-900 mb-4 pb-2 border-b-2 border-gray-200">Units</h2>\n'
            panel += f'<ol class="pl-6">\n'
            for ut in m["unit_titles"]:
                panel += f'<li class="my-1 text-sm">{escape(ut)}</li>\n'
            panel += f'</ol></div>\n'

        panel += f'</div>\n'  # close grade panel
        grade_panels += panel

    # --- Cross-grade summary ---
    summary_html = '<div id="summary" class="grade-panel" style="display:none;">\n'
    summary_html += '<div class="bg-white rounded-lg p-6 mb-5 shadow-sm"><h2 class="text-lg font-semibold text-indigo-900 mb-4 pb-2 border-b-2 border-gray-200">Cross-Grade Summary</h2>\n'

    summary_html += '<div class="overflow-x-auto">\n'
    summary_html += f'<table {TABLE}><thead><tr>'
    summary_html += f'<th {TH}>Grade</th><th {TH}>Articles</th><th {TH}>Questions</th><th {TH}>Units</th>'
    summary_html += f'<th {TH}>Literary %</th><th {TH}>Synopsis %</th><th {TH}>Excerpt %</th><th {TH}>A %</th><th {TH}>B %</th><th {TH}>C %</th><th {TH}>D %</th>'
    summary_html += f'<th {TH}>Flagged</th></tr></thead><tbody>\n'

    for grade in sorted(all_grades.keys()):
        m = all_grades[grade]
        a_pct = (m["answer_distribution"].get(0, 0) / m["answer_total"] * 100) if m["answer_total"] else 0
        b_pct = (m["answer_distribution"].get(1, 0) / m["answer_total"] * 100) if m["answer_total"] else 0
        c_pct = (m["answer_distribution"].get(2, 0) / m["answer_total"] * 100) if m["answer_total"] else 0
        d_pct = (m["answer_distribution"].get(3, 0) / m["answer_total"] * 100) if m["answer_total"] else 0

        # Synopsis percentage for grades 3-8
        if m["total_literary_unit_assessments"] > 0:
            syn_n = m["originality_counts"].get("synopsis", 0)
            syn_pct_val = syn_n / m["total_literary_unit_assessments"] * 100
            syn_cell = f'{syn_pct_val:.0f}%'
        else:
            syn_cell = "N/A"

        summary_html += f'<tr class="hover:bg-gray-50"><td {TD}><strong>Grade {grade}</strong></td>'
        summary_html += f'<td {TD}>{m["total_articles"]}</td><td {TD}>{m["total_questions"]}</td><td {TD}>{m["num_units"]}</td>'
        summary_html += f'<td {TD}>{m["literary_pct"]:.0f}%</td><td {TD}>{syn_cell}</td><td {TD}>{m["excerpt_pct"]:.0f}%</td>'
        summary_html += f'<td {TD}>{a_pct:.0f}%</td><td {TD}>{b_pct:.0f}%</td><td {TD}>{c_pct:.0f}%</td><td {TD}>{d_pct:.0f}%</td>'
        summary_html += f'<td {TD}>{len(m["flagged_items"])}</td></tr>\n'

    summary_html += '</tbody></table></div></div>\n'

    # Curriculum Plan vs Reality Summary (grades 9-12)
    has_plan_data = any(all_grades[g].get("plan_crossref") for g in range(9, 13) if g in all_grades)
    if has_plan_data:
        summary_html += '<div class="bg-white rounded-lg p-6 mb-5 shadow-sm"><h2 class="text-lg font-semibold text-indigo-900 mb-4 pb-2 border-b-2 border-gray-200">Curriculum Plan vs. QTI Reality (Grades 9–12)</h2>\n'
        summary_html += '<p class="text-xs text-gray-500 mb-3 italic">The high school curriculum plan specifies a comprehensive, well-structured program. '
        summary_html += 'This table compares the plan to what is actually implemented in the QTI assessment data.</p>\n'
        summary_html += '<div class="overflow-x-auto">\n'
        summary_html += f'<table {TABLE}><thead><tr>'
        summary_html += f'<th {TH}>Grade</th><th {TH}>Planned Texts</th><th {TH}>QTI Articles</th>'
        summary_html += f'<th {TH}>Matched</th><th {TH}>Match %</th><th {TH}>Missing</th><th {TH}>Custom Needed</th>'
        summary_html += f'<th {TH}>Extra in QTI</th><th {TH}>Planned Units</th><th {TH}>QTI Units</th>'
        summary_html += f'<th {TH}>Planned Lexile</th></tr></thead><tbody>\n'

        for grade in range(9, 13):
            if grade not in all_grades:
                continue
            m = all_grades[grade]
            xr = m.get("plan_crossref")
            cp = m.get("curriculum_plan")
            pl = PLANNED_LEXILE.get(grade, {})
            if not xr or not cp:
                continue

            match_tw = "text-green-600" if xr["match_rate"] > 50 else ("text-pink-600" if xr["match_rate"] > 20 else "text-red-600")
            unit_tw = "text-red-600" if m["num_units"] <= 1 else "text-green-600"
            summary_html += f'<tr class="hover:bg-gray-50"><td {TD}><strong>Grade {grade}</strong></td>'
            summary_html += f'<td {TD}>{cp["total"]}</td><td {TD}>{m["total_articles"]}</td>'
            summary_html += f'<td {TD}>{len(xr["found_in_qti"])}</td>'
            summary_html += f'<td class="px-3 py-2 border-b border-gray-200 {match_tw} font-semibold">{xr["match_rate"]:.0f}%</td>'
            summary_html += f'<td class="px-3 py-2 border-b border-gray-200 text-red-600">{len(xr["missing_from_qti"])}</td>'
            summary_html += f'<td class="px-3 py-2 border-b border-gray-200 text-pink-500">{len(xr["custom_not_yet_written"])}</td>'
            summary_html += f'<td {TD}>{len(xr["extra_in_qti"])}</td>'
            summary_html += f'<td {TD}>{len(cp["units"])}</td>'
            summary_html += f'<td class="px-3 py-2 border-b border-gray-200 {unit_tw}">{m["num_units"]}</td>'
            summary_html += f'<td {TD}>{pl.get("range", "?")}</td></tr>\n'

        summary_html += '</tbody></table></div></div>\n'

    # Cross-grade answer bias chart
    summary_html += '<div class="bg-white rounded-lg p-6 mb-5 shadow-sm"><h2 class="text-lg font-semibold text-indigo-900 mb-4 pb-2 border-b-2 border-gray-200">Answer Distribution Across Grades</h2>\n'
    summary_html += '<p class="text-xs text-gray-500 mb-3 italic">Expected: 25% per answer. Observed: systematic A/B bias across all grades.</p>\n'
    for grade in sorted(all_grades.keys()):
        m = all_grades[grade]
        summary_html += f'<div class="my-3"><strong>Grade {grade}</strong><div class="flex gap-1 mt-1">'
        for i, (label, bg_cls) in enumerate(zip(["A", "B", "C", "D"], ["bg-blue-400", "bg-emerald-400", "bg-cyan-400", "bg-red-400"])):
            pct = (m["answer_distribution"].get(i, 0) / m["answer_total"] * 100) if m["answer_total"] else 0
            summary_html += (
                f'<div class="flex-1 text-center">'
                f'<div class="rounded {bg_cls}" style="height:{max(pct * 2, 2):.0f}px;"></div>'
                f'<div class="text-[11px] mt-0.5">{label}: {pct:.0f}%</div></div>'
            )
        summary_html += '</div></div>\n'
    summary_html += '</div>\n'

    # Global weaknesses
    summary_html += '<div class="bg-white rounded-lg p-6 mb-5 shadow-sm"><h2 class="text-lg font-semibold text-indigo-900 mb-4 pb-2 border-b-2 border-gray-200">Systemic Weaknesses (All Grades)</h2>\n'
    summary_html += '<ul class="list-none">\n'
    global_weaknesses = [
        "Answer key bias is consistent across ALL grades — correct answers cluster in positions A and B, "
        "with C and D almost never correct. This is a systemic design flaw that compromises assessment integrity.",

        "Grades 3–8 literary units are overwhelmingly curriculum-written synopses, retellings, and educational "
        "articles ABOUT literary works rather than the works themselves. Students read about what happens "
        "in poems, stories, and plays instead of reading the actual texts. Only a small fraction of assessments "
        "present genuine original literary text. This fundamentally undermines CCSS standards requiring students "
        "to read and comprehend literature (RL standards).",

        "Grades 9–12 lack thematic unit organization entirely. All content is dumped into a single "
        "'Reading Exercises' unit, unlike the well-structured units in grades 3–8.",

        "Grades 9–12 have no Lexile data (all set to 0), making text complexity verification impossible. "
        "The curriculum plan specifies scaffolded Lexile progressions (e.g., Grade 9: 1050L–1260L, "
        "Grade 12: 1250L–1420L) but the QTI data does not implement these ranges.",

        "MAJOR GAP between curriculum plan and QTI reality for grades 9–12: The curriculum plan specifies "
        "well-organized thematic units (6–7 per grade), scaffolded Lexile progressions, 130–135 carefully "
        "selected texts with custom informational articles to be commissioned — but the QTI data implements "
        "none of this structure. All content is in a single undifferentiated unit with no Lexile data, "
        "no thematic organization, and a very different text selection from what the plan specifies.",
    ]
    for w in global_weaknesses:
        summary_html += f'<li class="my-2.5 pl-2 text-red-900 leading-normal before:content-[\'\\2717_\'] before:font-bold before:text-red-600">{escape(w)}</li>\n'
    summary_html += '</ul></div>\n'

    summary_html += '</div>\n'

    # --- Context Analysis Tab ---
    context_html = '<div id="context" class="grade-panel" style="display:none;">\n'
    context_html += '<div class="bg-white rounded-lg p-6 mb-5 shadow-sm">\n'
    context_html += '<h2 class="text-lg font-semibold text-indigo-900 mb-4 pb-2 border-b-2 border-gray-200">Contextualization Analysis (Grades 9–12)</h2>\n'
    context_html += '<p class="text-sm text-gray-600 mb-4">This analysis identifies texts that require a contextual introduction paragraph '
    context_html += 'before students begin reading. Without proper framing, students may misunderstand the text, miss critical context, '
    context_html += 'or be unable to engage meaningfully with the material.</p>\n'

    if context_data:
        # Summary stats
        total_flagged = sum(len(v) for v in context_data.values())
        has_context_count = sum(1 for v in context_data.values() for e in v if e["has_context"])
        no_context_count = total_flagged - has_context_count

        context_html += '<div class="grid grid-cols-[repeat(auto-fit,minmax(140px,1fr))] gap-4 mb-6">\n'
        context_html += f'<div class="text-center p-4 bg-gray-50 rounded-lg"><div class="text-3xl font-bold text-blue-600">{total_flagged}</div><div class="text-xs text-gray-500 mt-1">Texts Needing Context</div></div>\n'
        context_html += f'<div class="text-center p-4 bg-gray-50 rounded-lg"><div class="text-3xl font-bold text-red-500">{no_context_count}</div><div class="text-xs text-gray-500 mt-1">Missing Context</div></div>\n'
        context_html += f'<div class="text-center p-4 bg-gray-50 rounded-lg"><div class="text-3xl font-bold text-emerald-500">{has_context_count}</div><div class="text-xs text-gray-500 mt-1">Has Some Context</div></div>\n'
        context_html += '</div>\n'
        context_html += '</div>\n'

        # Category legend
        context_html += '<div class="bg-white rounded-lg p-6 mb-5 shadow-sm">\n'
        context_html += '<h3 class="text-base font-semibold text-indigo-900 mb-3">Categories</h3>\n'
        cat_colors = {
            "satire": ("bg-purple-100 text-purple-800", "bg-purple-400"),
            "ancient": ("bg-sky-100 text-sky-800", "bg-sky-400"),
            "mid-excerpt": ("bg-blue-100 text-blue-800", "bg-blue-400"),
            "primary-source": ("bg-teal-100 text-teal-800", "bg-teal-400"),
            "philosophical": ("bg-rose-100 text-rose-800", "bg-rose-400"),
        }
        context_html += '<div class="flex flex-wrap gap-3 mb-2">\n'
        for cat_key, cat_label in CATEGORY_LABELS.items():
            badge_cls = cat_colors.get(cat_key, ("bg-gray-100 text-gray-800", "bg-gray-500"))[0]
            context_html += f'<span class="text-xs px-2.5 py-1 rounded-full font-medium {badge_cls}">{escape(cat_label)}</span>\n'
        context_html += '</div>\n'
        context_html += '</div>\n'

        # Per-grade tables
        for grade in sorted(context_data.keys()):
            entries = context_data[grade]
            if not entries:
                continue

            grade_no_ctx = sum(1 for e in entries if not e["has_context"])
            grade_has_ctx = sum(1 for e in entries if e["has_context"])

            context_html += f'<div class="bg-white rounded-lg p-6 mb-5 shadow-sm">\n'
            context_html += f'<h2 class="text-lg font-semibold text-indigo-900 mb-2 pb-2 border-b-2 border-gray-200">Grade {grade}</h2>\n'
            context_html += f'<p class="text-xs text-gray-500 mb-4">{len(entries)} texts identified &mdash; '
            context_html += f'<span class="text-red-600 font-semibold">{grade_no_ctx} missing context</span>'
            if grade_has_ctx:
                context_html += f', <span class="text-emerald-600 font-semibold">{grade_has_ctx} have some context</span>'
            context_html += '</p>\n'

            context_html += f'<table {TABLE}>\n'
            context_html += f'<thead><tr><th {TH}>Text</th><th {TH}>Category</th><th {TH}>Context?</th><th {TH}>Why Context Is Needed</th></tr></thead>\n'
            context_html += '<tbody>\n'

            for entry in entries:
                badge_cls = cat_colors.get(entry["category"], ("bg-gray-100 text-gray-800", "bg-gray-500"))[0]
                cat_short = CATEGORY_LABELS.get(entry["category"], entry["category"]).split(" — ")[0]

                if entry["has_context"]:
                    status = f'<span class="text-xs px-2 py-0.5 rounded-full bg-emerald-100 text-emerald-700 font-medium">Partial</span>'
                    status += f'<div class="text-[11px] text-gray-500 mt-1">{escape(entry["context_note"])}</div>'
                else:
                    status = '<span class="text-xs px-2 py-0.5 rounded-full bg-red-100 text-red-700 font-medium">None</span>'

                row_bg = "" if not entry["has_context"] else ""
                context_html += f'<tr class="hover:bg-gray-50 align-top">'
                context_html += f'<td {TD}><strong class="text-sm">{escape(entry["title"][:80])}</strong>'
                # Show first line of actual text
                preview = entry["first_200"][:120].strip()
                if preview:
                    context_html += f'<div class="text-[11px] text-gray-400 mt-1 italic">Starts: {escape(preview)}...</div>'
                context_html += f'</td>'
                context_html += f'<td {TD}><span class="text-[11px] px-2 py-0.5 rounded-full font-medium {badge_cls}">{escape(cat_short)}</span></td>'
                context_html += f'<td {TD}>{status}</td>'
                context_html += f'<td {TD}><span class="text-sm">{escape(entry["reason"])}</span></td>'
                context_html += f'</tr>\n'

            context_html += '</tbody></table>\n'
            context_html += '</div>\n'
    else:
        context_html += '<p class="text-gray-500 italic">Context analysis not available (grades 9–12 data required).</p>\n'
        context_html += '</div>\n'

    context_html += '</div>\n'

    # Assemble full HTML
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Alpha Read Analysis Report</title>
<style>
{open(os.path.join(os.path.dirname(__file__), 'output.css')).read()}
/* Minimal styles for things Tailwind can't handle */
.tab-bar {{
    position: sticky;
    top: 0;
    z-index: 100;
}}
.tab-btn {{
    border: none;
    background: none;
}}
</style>
</head>
<body class="bg-slate-100 text-gray-900 leading-relaxed font-sans">
<div class="bg-gradient-to-br from-indigo-900 to-blue-500 text-white px-8 py-6 shadow-lg">
<h1 class="text-2xl font-semibold">Alpha Read — Curriculum Analysis Report</h1>
<p class="text-sm opacity-85 mt-1">Grades 3–12 Reading Plans &bull; Updated {__import__('datetime').datetime.now().strftime('%B %d, %Y')}</p>
</div>
<div class="tab-bar flex px-6 bg-white border-b-2 border-gray-200 overflow-x-auto shadow-sm">
<button class="tab-btn px-4 py-3 border-b-[3px] border-transparent text-sm font-medium text-gray-500 hover:text-blue-600 hover:bg-blue-50 cursor-pointer whitespace-nowrap transition-all" onclick="showGrade('summary', this)">Summary</button>
{grade_tabs}
<button class="tab-btn px-4 py-3 border-b-[3px] border-transparent text-sm font-medium text-gray-500 hover:text-blue-600 hover:bg-blue-50 cursor-pointer whitespace-nowrap transition-all" onclick="showGrade('context', this)">Context</button>
</div>
<div class="max-w-[1100px] mx-auto p-6">
{summary_html}
{grade_panels}
{context_html}
</div>
<script>
function showGrade(grade, btn) {{
    document.querySelectorAll('.grade-panel').forEach(p => p.style.display = 'none');
    document.querySelectorAll('.tab-btn').forEach(b => {{
        b.classList.remove('active', 'text-blue-600', 'border-blue-600', 'font-semibold');
    }});
    const panel = document.getElementById(grade === 'summary' || grade === 'context' ? grade : 'grade-' + grade);
    if (panel) panel.style.display = 'block';
    if (btn) {{
        btn.classList.add('active', 'text-blue-600', 'border-blue-600', 'font-semibold');
    }}
    window.scrollTo(0, 0);
}}
</script>
</body>
</html>"""

    return html


# ============================================================================
# MAIN
# ============================================================================

def main():
    print("Alpha Read Analysis Report Generator")
    print("=" * 50)

    # Parse curriculum plan (grades 9-12)
    print("Parsing curriculum plan XLSX...")
    curriculum_plan = parse_curriculum_plan()
    if curriculum_plan:
        for g, p in sorted(curriculum_plan.items()):
            print(f"  Plan Grade {g}: {p['total']} texts ({p['custom_count']} custom), {len(p['units'])} units")
    else:
        print("  (No curriculum plan data available)")

    all_grades = {}
    for grade in range(3, 13):
        print(f"Analyzing grade {grade}...")
        data = load_grade_data(grade)
        all_grades[grade] = analyze_grade(grade, data)
        m = all_grades[grade]

        # Cross-reference with curriculum plan for grades 9-12
        if grade in curriculum_plan:
            assessments = get_all_assessments(data)
            xref = cross_reference_plan(curriculum_plan[grade]["texts"], assessments, grade)
            m["curriculum_plan"] = curriculum_plan[grade]
            m["plan_crossref"] = xref
            print(f"  → {m['total_articles']} articles, {m['total_questions']} questions, "
                  f"{m['literary_pct']:.0f}% literary, {m['excerpt_pct']:.0f}% excerpts, "
                  f"{len(m['flagged_items'])} flagged | "
                  f"Plan match: {xref['match_rate']:.0f}% ({len(xref['found_in_qti'])}/{curriculum_plan[grade]['total']})")
        else:
            m["curriculum_plan"] = None
            m["plan_crossref"] = None
            print(f"  → {m['total_articles']} articles, {m['total_questions']} questions, "
                  f"{m['literary_pct']:.0f}% literary, {m['excerpt_pct']:.0f}% excerpts, "
                  f"{len(m['flagged_items'])} flagged")

    # Contextualization analysis for grades 9-12
    print("\nRunning contextualization analysis (grades 9-12)...")
    context_data = {}
    for grade in range(9, 13):
        data = load_grade_data(grade)
        ctx = analyze_contextualization(grade, data)
        context_data[grade] = ctx
        no_ctx = sum(1 for e in ctx if not e["has_context"])
        print(f"  Grade {grade}: {len(ctx)} texts need context, {no_ctx} missing it")

    print("\nGenerating HTML report...")
    html = generate_html(all_grades, context_data=context_data)

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"Report saved to: {OUTPUT_PATH}")
    print(f"File size: {os.path.getsize(OUTPUT_PATH) / 1024:.0f} KB")


if __name__ == "__main__":
    main()
